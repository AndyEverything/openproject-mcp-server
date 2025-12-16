#!/usr/bin/env python3
"""
OpenProject MCP Server

A Model Context Protocol (MCP) server that provides integration with OpenProject API v3.
Supports project management, work package tracking, and task creation through a
standardized interface.
"""

import os
import json
import logging
from typing import Dict, List, Optional, Any
from datetime import datetime
import asyncio
import aiohttp
from urllib.parse import quote
import base64
import ssl
from dotenv import load_dotenv
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available. Install with: pip install openpyxl")

from mcp.server import Server
from mcp.types import (
    Tool,
    TextContent,
)

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(
    level=os.getenv("LOG_LEVEL", "INFO"),
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)

# Version information
__version__ = "1.0.0"
__author__ = "Your Name"
__license__ = "MIT"

# Default page size for API requests (to handle large datasets)
DEFAULT_PAGE_SIZE = 50000


class OpenProjectClient:
    """Client for the OpenProject API v3 with optional proxy support"""

    def __init__(self, base_url: str, api_key: str, proxy: Optional[str] = None):
        """
        Initialize the OpenProject client.

        Args:
            base_url: The base URL of the OpenProject instance
            api_key: API key for authentication
            proxy: Optional HTTP proxy URL
        """
        self.base_url = base_url.rstrip("/")
        self.api_key = api_key
        self.proxy = proxy

        # Setup headers with Basic Auth
        self.headers = {
            "Authorization": f"Basic {self._encode_api_key()}",
            "Content-Type": "application/json",
            "Accept": "application/json",
            "User-Agent": f"OpenProject-MCP/{__version__}",
        }

        logger.info(f"OpenProject Client initialized for: {self.base_url}")
        if self.proxy:
            logger.info(f"Using proxy: {self.proxy}")

    def _encode_api_key(self) -> str:
        """Encode API key for Basic Auth"""
        credentials = f"apikey:{self.api_key}"
        return base64.b64encode(credentials.encode()).decode()

    async def _request(
        self, method: str, endpoint: str, data: Optional[Dict] = None
    ) -> Dict:
        """
        Execute an API request.

        Args:
            method: HTTP method (GET, POST, etc.)
            endpoint: API endpoint path
            data: Optional request body data

        Returns:
            Dict: Response data from the API

        Raises:
            Exception: If the request fails
        """
        url = f"{self.base_url}/api/v3{endpoint}"

        logger.debug(f"API Request: {method} {url}")
        if data:
            logger.debug(f"Request body: {json.dumps(data, indent=2)}")

        # Configure SSL and timeout
        # Check if SSL verification should be disabled (useful for self-signed certificates)
        verify_ssl = os.getenv("OPENPROJECT_VERIFY_SSL", "true").lower() == "true"

        if verify_ssl:
            ssl_context = ssl.create_default_context()
        else:
            logger.warning("SSL certificate verification is DISABLED. This is insecure for production use.")
            ssl_context = ssl.create_default_context()
            ssl_context.check_hostname = False
            ssl_context.verify_mode = ssl.CERT_NONE

        connector = aiohttp.TCPConnector(ssl=ssl_context)
        timeout = aiohttp.ClientTimeout(total=30)

        async with aiohttp.ClientSession(
            connector=connector, timeout=timeout
        ) as session:
            try:
                # Build request parameters
                request_params = {
                    "method": method,
                    "url": url,
                    "headers": self.headers,
                    "json": data,
                }

                # Add proxy if configured
                if self.proxy:
                    request_params["proxy"] = self.proxy

                async with session.request(**request_params) as response:
                    response_text = await response.text()

                    logger.debug(f"Response status: {response.status}")

                    # Parse response
                    try:
                        response_json = (
                            json.loads(response_text) if response_text else {}
                        )
                    except json.JSONDecodeError:
                        logger.error(f"Invalid JSON response: {response_text[:200]}...")
                        response_json = {}

                    # Handle errors
                    if response.status >= 400:
                        error_msg = self._format_error_message(
                            response.status, response_text
                        )
                        raise Exception(error_msg)

                    return response_json

            except aiohttp.ClientError as e:
                logger.error(f"Network error: {str(e)}")
                raise Exception(f"Network error accessing {url}: {str(e)}")

    def _format_error_message(self, status: int, response_text: str) -> str:
        """Format error message based on HTTP status code"""
        base_msg = f"API Error {status}: {response_text}"

        error_hints = {
            401: "Authentication failed. Please check your API key.",
            403: "Access denied. The user lacks required permissions.",
            404: "Resource not found. Please verify the URL and resource exists.",
            407: "Proxy authentication required.",
            500: "Internal server error. Please try again later.",
            502: "Bad gateway. The server or proxy is not responding correctly.",
            503: "Service unavailable. The server might be under maintenance.",
        }

        if status in error_hints:
            base_msg += f"\n\n{error_hints[status]}"

        return base_msg

    async def test_connection(self) -> Dict:
        """Test the API connection and authentication"""
        logger.info("Testing API connection...")
        return await self._request("GET", "")

    async def get_projects(self, filters: Optional[str] = None, page_size: int = DEFAULT_PAGE_SIZE) -> Dict:
        """
        Retrieve all projects.

        Args:
            filters: Optional JSON-encoded filter string
            page_size: Number of results per page (default: DEFAULT_PAGE_SIZE)

        Returns:
            Dict: API response containing projects
        """
        endpoint = "/projects"
        query_params = []
        
        if filters:
            encoded_filters = quote(filters)
            query_params.append(f"filters={encoded_filters}")
        
        # Add pageSize parameter to fetch up to DEFAULT_PAGE_SIZE projects
        query_params.append(f"pageSize={page_size}")
        
        if query_params:
            endpoint += "?" + "&".join(query_params)

        result = await self._request("GET", endpoint)

        # Ensure proper response structure
        if "_embedded" not in result:
            result["_embedded"] = {"elements": []}
        elif "elements" not in result.get("_embedded", {}):
            result["_embedded"]["elements"] = []

        return result

    async def get_work_packages(
        self,
        project_id: Optional[int] = None,
        filters: Optional[str] = None,
        offset: Optional[int] = None,
        page_size: Optional[int] = DEFAULT_PAGE_SIZE,
    ) -> Dict:
        """
        Retrieve work packages.

        Args:
            project_id: Optional project ID to filter by
            filters: Optional JSON-encoded filter string
            offset: Optional starting index for pagination
            page_size: Optional number of results per page (default: DEFAULT_PAGE_SIZE)

        Returns:
            Dict: API response containing work packages
        """
        if project_id:
            endpoint = f"/projects/{project_id}/work_packages"
        else:
            endpoint = "/work_packages"

        # Build query parameters
        query_params = []
        if filters:
            encoded_filters = quote(filters)
            query_params.append(f"filters={encoded_filters}")
        if offset is not None:
            query_params.append(f"offset={offset}")
        if page_size is not None:
            query_params.append(f"pageSize={page_size}")

        if query_params:
            endpoint += "?" + "&".join(query_params)

        result = await self._request("GET", endpoint)

        # Ensure proper response structure
        if "_embedded" not in result:
            result["_embedded"] = {"elements": []}
        elif "elements" not in result.get("_embedded", {}):
            result["_embedded"]["elements"] = []

        return result

    async def create_work_package(self, data: Dict) -> Dict:
        """
        Create a new work package.

        Args:
            data: Work package data including project, subject, type, etc.

        Returns:
            Dict: Created work package data
        """
        # Prepare initial payload for form
        form_payload = {"_links": {}}

        # Set required links
        if "project" in data:
            form_payload["_links"]["project"] = {
                "href": f"/api/v3/projects/{data['project']}"
            }
        if "type" in data:
            form_payload["_links"]["type"] = {"href": f"/api/v3/types/{data['type']}"}

        # Set subject if provided
        if "subject" in data:
            form_payload["subject"] = data["subject"]

        # Get form with initial payload
        form = await self._request("POST", "/work_packages/form", form_payload)

        # Use form payload and add additional fields
        payload = form.get("payload", form_payload)
        payload["lockVersion"] = form.get("lockVersion", 0)

        # Add optional fields
        if "description" in data:
            payload["description"] = {"raw": data["description"]}
        if "priority_id" in data:
            if "_links" not in payload:
                payload["_links"] = {}
            payload["_links"]["priority"] = {
                "href": f"/api/v3/priorities/{data['priority_id']}"
            }
        if "assignee_id" in data:
            if "_links" not in payload:
                payload["_links"] = {}
            payload["_links"]["assignee"] = {
                "href": f"/api/v3/users/{data['assignee_id']}"
            }

        # Add date fields (ISO 8601 format: YYYY-MM-DD)
        if "startDate" in data:
            payload["startDate"] = data["startDate"]
        if "dueDate" in data:
            payload["dueDate"] = data["dueDate"]
        if "date" in data:
            payload["date"] = data["date"]

        # Create work package
        return await self._request("POST", "/work_packages", payload)

    async def get_types(self, project_id: Optional[int] = None) -> Dict:
        """
        Retrieve available work package types.

        Args:
            project_id: Optional project ID to filter types by

        Returns:
            Dict: API response containing types
        """
        if project_id:
            endpoint = f"/projects/{project_id}/types?pageSize={DEFAULT_PAGE_SIZE}"
        else:
            endpoint = f"/types?pageSize={DEFAULT_PAGE_SIZE}"

        result = await self._request("GET", endpoint)

        # Ensure proper response structure
        if "_embedded" not in result:
            result["_embedded"] = {"elements": []}
        elif "elements" not in result.get("_embedded", {}):
            result["_embedded"]["elements"] = []

        return result

    async def get_users(self, filters: Optional[str] = None, page_size: int = DEFAULT_PAGE_SIZE) -> Dict:
        """
        Retrieve users.

        Args:
            filters: Optional JSON-encoded filter string
            page_size: Number of results per page (default: DEFAULT_PAGE_SIZE)

        Returns:
            Dict: API response containing users
        """
        endpoint = "/users"
        query_params = []
        
        if filters:
            encoded_filters = quote(filters)
            query_params.append(f"filters={encoded_filters}")
        
        # Add pageSize parameter to fetch more than default 50 users
        query_params.append(f"pageSize={page_size}")
        
        if query_params:
            endpoint += "?" + "&".join(query_params)

        result = await self._request("GET", endpoint)

        # Ensure proper response structure
        if "_embedded" not in result:
            result["_embedded"] = {"elements": []}
        elif "elements" not in result.get("_embedded", {}):
            result["_embedded"]["elements"] = []

        return result

    async def get_user(self, user_id: int) -> Dict:
        """
        Retrieve a specific user by ID.

        Args:
            user_id: The user ID

        Returns:
            Dict: User data
        """
        return await self._request("GET", f"/users/{user_id}")

    async def create_comment(self, work_package_id: int, comment: str) -> Dict:
        """
        Create a comment on a work package.

        Args:
            work_package_id: The work package ID
            comment: The comment text

        Returns:
            Dict: Created comment data
        """
        payload = {
            "comment": {
                "raw": comment
            }
        }
        return await self._request("POST", f"/work_packages/{work_package_id}/activities", payload)

    async def get_comment(self, work_package_id: int) -> Dict:
        """
        Get comments for a work package.

        Args:
            work_package_id: The work package ID

        Returns:
            Dict: Comments data
        """
        result = await self._request("GET", f"/work_packages/{work_package_id}/activities?pageSize={DEFAULT_PAGE_SIZE}")
        
        # Ensure proper response structure
        if "_embedded" not in result:
            result["_embedded"] = {"elements": []}
        elif "elements" not in result.get("_embedded", {}):
            result["_embedded"]["elements"] = []
        
        return result

    async def update_comment(self, comment_id: int, comment: str) -> Dict:
        """
        Update a comment.

        Args:
            comment_id: The comment ID
            comment: The new comment text

        Returns:
            Dict: Updated comment data
        """
        payload = {
            "comment": {
                "raw": comment
            }
        }
        return await self._request("PATCH", f"/activities/{comment_id}", payload)

    async def get_queries(self, project_id: Optional[int] = None) -> Dict:
        """
        Get available queries (views) for a project or globally.

        Args:
            project_id: Optional project ID to get project-specific queries

        Returns:
            Dict: Queries data
        """
        endpoint = f"/queries?pageSize={DEFAULT_PAGE_SIZE}"
        
        result = await self._request("GET", endpoint)
        
        # Ensure proper response structure
        if "_embedded" not in result:
            result["_embedded"] = {"elements": []}
        elif "elements" not in result.get("_embedded", {}):
            result["_embedded"]["elements"] = []
        
        # Client-side filtering by project if needed
        if project_id:
            all_queries = result.get("_embedded", {}).get("elements", [])
            filtered_queries = []
            for query in all_queries:
                # Check if query belongs to the project
                query_project_link = query.get("_links", {}).get("project", {}).get("href", "")
                if query_project_link and (query_project_link.endswith(f"/{project_id}") or query_project_link.endswith(f"/projects/{project_id}")):
                    filtered_queries.append(query)
                # Also check if there's no project link (might be global or user query)
                elif not query_project_link and query.get("_links", {}).get("project") is None:
                    # Skip queries without project association when filtering by project
                    pass
            result["_embedded"]["elements"] = filtered_queries
        
        return result

    async def get_query(self, query_id: int) -> Dict:
        """
        Get a specific query (view) by ID.

        Args:
            query_id: The query ID

        Returns:
            Dict: Query data with filters and columns
        """
        return await self._request("GET", f"/queries/{query_id}")

    async def get_query_results(self, query_id: int) -> Dict:
        """
        Get work packages results for a specific query.

        Args:
            query_id: The query ID

        Returns:
            Dict: Work packages matching the query
        """
        # First get the query to extract filters and configuration
        query = await self._request("GET", f"/queries/{query_id}")
        
        # Try to get the results href from the query's _links
        if "_links" in query and "results" in query["_links"]:
            results_href = query["_links"]["results"]["href"]
            # Extract the path from the href (remove /api/v3 prefix if present)
            if results_href.startswith("/api/v3"):
                results_path = results_href[7:]  # Remove '/api/v3'
            else:
                results_path = results_href
            
            result = await self._request("GET", results_path)
        else:
            # Fallback: fetch all work packages and filter by project if available
            endpoint = "/work_packages"
            
            # Try to get project from query
            if "_links" in query and "project" in query["_links"]:
                project_href = query["_links"]["project"]["href"]
                project_id = project_href.split("/")[-1]
                endpoint = f"/projects/{project_id}/work_packages"
            
            result = await self._request("GET", endpoint)
        
        # Ensure proper response structure
        if "_embedded" not in result:
            result["_embedded"] = {"elements": []}
        elif "elements" not in result.get("_embedded", {}):
            result["_embedded"]["elements"] = []
        
        return result

    async def get_memberships(
        self, project_id: Optional[int] = None, user_id: Optional[int] = None
    ) -> Dict:
        """
        Retrieve memberships.

        Args:
            project_id: Optional project ID to filter memberships by project
            user_id: Optional user ID to filter memberships by user

        Returns:
            Dict: API response containing memberships
        """
        endpoint = "/memberships"
        query_params = []
        
        # Add page size for large datasets
        query_params.append(f"pageSize={DEFAULT_PAGE_SIZE}")

        # Use filters instead of path-based filtering for better compatibility
        filters = []
        if project_id:
            filters.append({"project": {"operator": "=", "values": [str(project_id)]}})
        if user_id:
            filters.append({"user": {"operator": "=", "values": [str(user_id)]}})

        if filters:
            filter_string = quote(json.dumps(filters))
            query_params.append(f"filters={filter_string}")
            
        if query_params:
            endpoint += "?" + "&".join(query_params)

        result = await self._request("GET", endpoint)

        # Ensure proper response structure
        if "_embedded" not in result:
            result["_embedded"] = {"elements": []}
        elif "elements" not in result.get("_embedded", {}):
            result["_embedded"]["elements"] = []

        return result

    async def get_statuses(self) -> Dict:
        """
        Retrieve available work package statuses.

        Returns:
            Dict: API response containing statuses
        """
        result = await self._request("GET", f"/statuses?pageSize={DEFAULT_PAGE_SIZE}")

        # Ensure proper response structure
        if "_embedded" not in result:
            result["_embedded"] = {"elements": []}
        elif "elements" not in result.get("_embedded", {}):
            result["_embedded"]["elements"] = []

        return result

    async def get_priorities(self) -> Dict:
        """
        Retrieve available work package priorities.

        Returns:
            Dict: API response containing priorities
        """
        result = await self._request("GET", f"/priorities?pageSize={DEFAULT_PAGE_SIZE}")

        # Ensure proper response structure
        if "_embedded" not in result:
            result["_embedded"] = {"elements": []}
        elif "elements" not in result.get("_embedded", {}):
            result["_embedded"]["elements"] = []

        return result

    async def get_work_package(self, work_package_id: int) -> Dict:
        """
        Retrieve a specific work package by ID.

        Args:
            work_package_id: The work package ID

        Returns:
            Dict: Work package data
        """
        return await self._request("GET", f"/work_packages/{work_package_id}")

    async def update_work_package(self, work_package_id: int, data: Dict) -> Dict:
        """
        Update an existing work package.

        Args:
            work_package_id: The work package ID
            data: Update data including fields to modify

        Returns:
            Dict: Updated work package data
        """
        # First get current work package to get lock version
        current_wp = await self.get_work_package(work_package_id)

        # Prepare payload with lock version
        payload = {"lockVersion": current_wp.get("lockVersion", 0)}

        # Add fields to update
        if "subject" in data:
            payload["subject"] = data["subject"]
        if "description" in data:
            payload["description"] = {"raw": data["description"]}
        if "type_id" in data:
            if "_links" not in payload:
                payload["_links"] = {}
            payload["_links"]["type"] = {"href": f"/api/v3/types/{data['type_id']}"}
        if "status_id" in data:
            if "_links" not in payload:
                payload["_links"] = {}
            payload["_links"]["status"] = {
                "href": f"/api/v3/statuses/{data['status_id']}"
            }
        if "priority_id" in data:
            if "_links" not in payload:
                payload["_links"] = {}
            payload["_links"]["priority"] = {
                "href": f"/api/v3/priorities/{data['priority_id']}"
            }
        if "assignee_id" in data:
            if "_links" not in payload:
                payload["_links"] = {}
            payload["_links"]["assignee"] = {
                "href": f"/api/v3/users/{data['assignee_id']}"
            }
        if "percentage_done" in data:
            payload["percentageDone"] = data["percentage_done"]

        # Add date fields (ISO 8601 format: YYYY-MM-DD)
        if "startDate" in data:
            payload["startDate"] = data["startDate"]
        if "dueDate" in data:
            payload["dueDate"] = data["dueDate"]
        if "date" in data:
            payload["date"] = data["date"]

        return await self._request(
            "PATCH", f"/work_packages/{work_package_id}", payload
        )

    async def delete_work_package(self, work_package_id: int) -> bool:
        """
        Delete a work package.

        Args:
            work_package_id: The work package ID

        Returns:
            bool: True if successful
        """
        await self._request("DELETE", f"/work_packages/{work_package_id}")
        return True

    async def get_time_entries(self, filters: Optional[str] = None) -> Dict:
        """
        Retrieve time entries.

        Args:
            filters: Optional JSON-encoded filter string

        Returns:
            Dict: API response containing time entries
        """
        endpoint = "/time_entries"
        query_params = []
        
        # Add page size for large datasets
        query_params.append(f"pageSize={DEFAULT_PAGE_SIZE}")
        
        if filters:
            encoded_filters = quote(filters)
            query_params.append(f"filters={encoded_filters}")
            
        if query_params:
            endpoint += "?" + "&".join(query_params)

        result = await self._request("GET", endpoint)

        # Ensure proper response structure
        if "_embedded" not in result:
            result["_embedded"] = {"elements": []}
        elif "elements" not in result.get("_embedded", {}):
            result["_embedded"]["elements"] = []

        return result

    async def create_time_entry(self, data: Dict) -> Dict:
        """
        Create a new time entry.

        Args:
            data: Time entry data including work package, hours, etc.

        Returns:
            Dict: Created time entry data
        """
        # Prepare payload
        payload = {}

        # Set required fields
        if "work_package_id" in data:
            payload["_links"] = {
                "workPackage": {
                    "href": f"/api/v3/work_packages/{data['work_package_id']}"
                }
            }
        if "hours" in data:
            payload["hours"] = f"PT{data['hours']}H"
        if "spent_on" in data:
            payload["spentOn"] = data["spent_on"]
        if "comment" in data:
            payload["comment"] = {"raw": data["comment"]}
        if "activity_id" in data:
            if "_links" not in payload:
                payload["_links"] = {}
            payload["_links"]["activity"] = {
                "href": f"/api/v3/time_entries/activities/{data['activity_id']}"
            }

        return await self._request("POST", "/time_entries", payload)

    async def update_time_entry(self, time_entry_id: int, data: Dict) -> Dict:
        """
        Update an existing time entry.

        Args:
            time_entry_id: The time entry ID
            data: Update data including fields to modify

        Returns:
            Dict: Updated time entry data
        """
        # First get current time entry to get lock version
        current_te = await self._request("GET", f"/time_entries/{time_entry_id}")

        # Prepare payload with lock version
        payload = {"lockVersion": current_te.get("lockVersion", 0)}

        # Add fields to update
        if "hours" in data:
            payload["hours"] = f"PT{data['hours']}H"
        if "spent_on" in data:
            payload["spentOn"] = data["spent_on"]
        if "comment" in data:
            payload["comment"] = {"raw": data["comment"]}
        if "activity_id" in data:
            if "_links" not in payload:
                payload["_links"] = {}
            payload["_links"]["activity"] = {
                "href": f"/api/v3/time_entries/activities/{data['activity_id']}"
            }

        return await self._request("PATCH", f"/time_entries/{time_entry_id}", payload)

    async def delete_time_entry(self, time_entry_id: int) -> bool:
        """
        Delete a time entry.

        Args:
            time_entry_id: The time entry ID

        Returns:
            bool: True if successful
        """
        await self._request("DELETE", f"/time_entries/{time_entry_id}")
        return True

    async def get_time_entry_activities(self) -> Dict:
        """
        Retrieve available time entry activities.

        Returns:
            Dict: API response containing activities
        """
        result = await self._request("GET", f"/time_entry_activities?pageSize={DEFAULT_PAGE_SIZE}")

        # Ensure proper response structure
        if "_embedded" not in result:
            result["_embedded"] = {"elements": []}
        elif "elements" not in result.get("_embedded", {}):
            result["_embedded"]["elements"] = []

        return result

    async def get_versions(self, project_id: Optional[int] = None) -> Dict:
        """
        Retrieve project versions.

        Args:
            project_id: Optional project ID to filter versions by project

        Returns:
            Dict: API response containing versions
        """
        if project_id:
            endpoint = f"/projects/{project_id}/versions?pageSize={DEFAULT_PAGE_SIZE}"
        else:
            endpoint = f"/versions?pageSize={DEFAULT_PAGE_SIZE}"

        result = await self._request("GET", endpoint)

        # Ensure proper response structure
        if "_embedded" not in result:
            result["_embedded"] = {"elements": []}
        elif "elements" not in result.get("_embedded", {}):
            result["_embedded"]["elements"] = []

        return result

    async def create_version(self, project_id: int, data: Dict) -> Dict:
        """
        Create a new project version.

        Args:
            project_id: The project ID
            data: Version data including name, description, etc.

        Returns:
            Dict: Created version data
        """
        # Prepare payload
        payload = {
            "_links": {"definingProject": {"href": f"/api/v3/projects/{project_id}"}}
        }

        # Set required fields
        if "name" in data:
            payload["name"] = data["name"]
        if "description" in data:
            payload["description"] = {"raw": data["description"]}
        if "start_date" in data:
            payload["startDate"] = data["start_date"]
        if "end_date" in data:
            payload["endDate"] = data["end_date"]
        if "status" in data:
            payload["status"] = data["status"]

        return await self._request("POST", "/versions", payload)

    async def check_permissions(self) -> Dict:
        """
        Check user permissions and capabilities.

        Returns:
            Dict: User information including permissions
        """
        try:
            # Get current user info which includes permissions
            return await self._request("GET", "/users/me")
        except Exception as e:
            logger.error(f"Failed to check permissions: {e}")
            return {}

    async def create_project(self, data: Dict) -> Dict:
        """
        Create a new project.

        Args:
            data: Project data including name, identifier, description, etc.

        Returns:
            Dict: Created project data
        """
        # Prepare payload
        payload = {}

        # Set required fields
        if "name" in data:
            payload["name"] = data["name"]
        if "identifier" in data:
            payload["identifier"] = data["identifier"]
        if "description" in data:
            payload["description"] = {"raw": data["description"]}
        if "public" in data:
            payload["public"] = data["public"]
        if "status" in data:
            payload["status"] = data["status"]
        if "parent_id" in data:
            if "_links" not in payload:
                payload["_links"] = {}
            payload["_links"]["parent"] = {
                "href": f"/api/v3/projects/{data['parent_id']}"
            }

        return await self._request("POST", "/projects", payload)

    async def update_project(self, project_id: int, data: Dict) -> Dict:
        """
        Update an existing project.

        Args:
            project_id: The project ID
            data: Update data including fields to modify

        Returns:
            Dict: Updated project data
        """
        # First get current project to get lock version if needed
        try:
            current_project = await self.get_project(project_id)
            lock_version = current_project.get("lockVersion", 0)
        except:
            lock_version = 0

        # Prepare payload with lock version
        payload = {"lockVersion": lock_version}

        # Add fields to update
        if "name" in data:
            payload["name"] = data["name"]
        if "identifier" in data:
            payload["identifier"] = data["identifier"]
        if "description" in data:
            payload["description"] = {"raw": data["description"]}
        if "public" in data:
            payload["public"] = data["public"]
        if "status" in data:
            payload["status"] = data["status"]
        if "parent_id" in data:
            if "_links" not in payload:
                payload["_links"] = {}
            payload["_links"]["parent"] = {
                "href": f"/api/v3/projects/{data['parent_id']}"
            }

        return await self._request("PATCH", f"/projects/{project_id}", payload)

    async def delete_project(self, project_id: int) -> bool:
        """
        Delete a project.

        Args:
            project_id: The project ID

        Returns:
            bool: True if successful
        """
        await self._request("DELETE", f"/projects/{project_id}")
        return True

    async def get_project(self, project_id: int) -> Dict:
        """
        Retrieve a specific project by ID.

        Args:
            project_id: The project ID

        Returns:
            Dict: Project data
        """
        return await self._request("GET", f"/projects/{project_id}")

    async def get_roles(self) -> Dict:
        """
        Retrieve available roles.

        Returns:
            Dict: API response containing roles
        """
        result = await self._request("GET", f"/roles?pageSize={DEFAULT_PAGE_SIZE}")

        # Ensure proper response structure
        if "_embedded" not in result:
            result["_embedded"] = {"elements": []}
        elif "elements" not in result.get("_embedded", {}):
            result["_embedded"]["elements"] = []

        return result

    async def get_role(self, role_id: int) -> Dict:
        """
        Retrieve a specific role by ID.

        Args:
            role_id: The role ID

        Returns:
            Dict: Role data
        """
        return await self._request("GET", f"/roles/{role_id}")

    async def create_membership(self, data: Dict) -> Dict:
        """
        Create a new membership.

        Args:
            data: Membership data including project, user/group, and roles

        Returns:
            Dict: Created membership data
        """
        # Prepare payload
        payload = {"_links": {}}

        # Set required fields
        if "project_id" in data:
            payload["_links"]["project"] = {
                "href": f"/api/v3/projects/{data['project_id']}"
            }
        if "user_id" in data:
            payload["_links"]["principal"] = {
                "href": f"/api/v3/users/{data['user_id']}"
            }
        elif "group_id" in data:
            payload["_links"]["principal"] = {
                "href": f"/api/v3/groups/{data['group_id']}"
            }
        if "role_ids" in data:
            payload["_links"]["roles"] = [
                {"href": f"/api/v3/roles/{role_id}"} for role_id in data["role_ids"]
            ]
        elif "role_id" in data:
            payload["_links"]["roles"] = [{"href": f"/api/v3/roles/{data['role_id']}"}]
        if "notification_message" in data:
            payload["notificationMessage"] = {"raw": data["notification_message"]}

        return await self._request("POST", "/memberships", payload)

    async def update_membership(self, membership_id: int, data: Dict) -> Dict:
        """
        Update an existing membership.

        Args:
            membership_id: The membership ID
            data: Update data including fields to modify

        Returns:
            Dict: Updated membership data
        """
        # First get current membership to get lock version if needed
        try:
            current_membership = await self.get_membership(membership_id)
            lock_version = current_membership.get("lockVersion", 0)
        except:
            lock_version = 0

        # Prepare payload with lock version
        payload = {"lockVersion": lock_version}

        # Add fields to update
        if "role_ids" in data:
            if "_links" not in payload:
                payload["_links"] = {}
            payload["_links"]["roles"] = [
                {"href": f"/api/v3/roles/{role_id}"} for role_id in data["role_ids"]
            ]
        elif "role_id" in data:
            if "_links" not in payload:
                payload["_links"] = {}
            payload["_links"]["roles"] = [{"href": f"/api/v3/roles/{data['role_id']}"}]
        if "notification_message" in data:
            payload["notificationMessage"] = {"raw": data["notification_message"]}

        return await self._request("PATCH", f"/memberships/{membership_id}", payload)

    async def delete_membership(self, membership_id: int) -> bool:
        """
        Delete a membership.

        Args:
            membership_id: The membership ID

        Returns:
            bool: True if successful
        """
        await self._request("DELETE", f"/memberships/{membership_id}")
        return True

    async def get_membership(self, membership_id: int) -> Dict:
        """
        Retrieve a specific membership by ID.

        Args:
            membership_id: The membership ID

        Returns:
            Dict: Membership data
        """
        return await self._request("GET", f"/memberships/{membership_id}")

    async def set_work_package_parent(
        self, work_package_id: int, parent_id: int
    ) -> Dict:
        """
        Set a parent for a work package (create parent-child relationship).

        Args:
            work_package_id: The work package ID to become a child
            parent_id: The work package ID to become the parent

        Returns:
            Dict: Updated work package data
        """
        # First get current work package to get lock version
        try:
            current_wp = await self.get_work_package(work_package_id)
            lock_version = current_wp.get("lockVersion", 0)
        except:
            lock_version = 0

        # Prepare payload with parent link
        payload = {
            "lockVersion": lock_version,
            "_links": {"parent": {"href": f"/api/v3/work_packages/{parent_id}"}},
        }

        return await self._request(
            "PATCH", f"/work_packages/{work_package_id}", payload
        )

    async def remove_work_package_parent(self, work_package_id: int) -> Dict:
        """
        Remove parent relationship from a work package (make it top-level).

        Args:
            work_package_id: The work package ID to remove parent from

        Returns:
            Dict: Updated work package data
        """
        # First get current work package to get lock version
        try:
            current_wp = await self.get_work_package(work_package_id)
            lock_version = current_wp.get("lockVersion", 0)
        except:
            lock_version = 0

        # Prepare payload with null parent link
        payload = {"lockVersion": lock_version, "_links": {"parent": None}}

        return await self._request(
            "PATCH", f"/work_packages/{work_package_id}", payload
        )

    async def list_work_package_children(
        self, parent_id: int, include_descendants: bool = False
    ) -> Dict:
        """
        List all child work packages of a parent.

        Args:
            parent_id: The parent work package ID
            include_descendants: If True, includes grandchildren and below

        Returns:
            Dict: API response containing child work packages
        """
        if include_descendants:
            # Use descendants filter to get all levels
            filters = json.dumps(
                [{"descendantsOf": {"operator": "=", "values": [str(parent_id)]}}]
            )
        else:
            # Use parent filter to get direct children only
            filters = json.dumps(
                [{"parent": {"operator": "=", "values": [str(parent_id)]}}]
            )

        endpoint = f"/work_packages?filters={quote(filters)}&pageSize={DEFAULT_PAGE_SIZE}"
        result = await self._request("GET", endpoint)

        # Ensure proper response structure
        if "_embedded" not in result:
            result["_embedded"] = {"elements": []}
        elif "elements" not in result.get("_embedded", {}):
            result["_embedded"]["elements"] = []

        return result

    async def create_work_package_relation(self, data: Dict) -> Dict:
        """
        Create a relationship between work packages.

        Args:
            data: Relation data including from_id, to_id, relation_type, lag

        Returns:
            Dict: Created relation data
        """
        # Prepare payload
        payload = {"_links": {}}

        # Set required fields
        if "from_id" in data:
            payload["_links"]["from"] = {
                "href": f"/api/v3/work_packages/{data['from_id']}"
            }
        if "to_id" in data:
            payload["_links"]["to"] = {"href": f"/api/v3/work_packages/{data['to_id']}"}
        if "relation_type" in data:
            payload["type"] = data["relation_type"]
        if "lag" in data:
            payload["lag"] = data["lag"]
        if "description" in data:
            payload["description"] = data["description"]

        return await self._request("POST", "/relations", payload)

    async def list_work_package_relations(self, filters: Optional[str] = None) -> Dict:
        """
        List work package relations.

        Args:
            filters: Optional JSON-encoded filter string

        Returns:
            Dict: API response containing relations
        """
        endpoint = "/relations"
        query_params = []
        
        # Add page size for large datasets
        query_params.append(f"pageSize={DEFAULT_PAGE_SIZE}")
        
        if filters:
            encoded_filters = quote(filters)
            query_params.append(f"filters={encoded_filters}")
            
        if query_params:
            endpoint += "?" + "&".join(query_params)

        result = await self._request("GET", endpoint)

        # Ensure proper response structure
        if "_embedded" not in result:
            result["_embedded"] = {"elements": []}
        elif "elements" not in result.get("_embedded", {}):
            result["_embedded"]["elements"] = []

        return result

    async def update_work_package_relation(self, relation_id: int, data: Dict) -> Dict:
        """
        Update an existing work package relation.

        Args:
            relation_id: The relation ID
            data: Update data including fields to modify

        Returns:
            Dict: Updated relation data
        """
        # First get current relation to get lock version if needed
        try:
            current_relation = await self.get_work_package_relation(relation_id)
            lock_version = current_relation.get("lockVersion", 0)
        except:
            lock_version = 0

        # Prepare payload with lock version
        payload = {"lockVersion": lock_version}

        # Add fields to update
        if "relation_type" in data:
            payload["type"] = data["relation_type"]
        if "lag" in data:
            payload["lag"] = data["lag"]
        if "description" in data:
            payload["description"] = data["description"]

        return await self._request("PATCH", f"/relations/{relation_id}", payload)

    async def delete_work_package_relation(self, relation_id: int) -> bool:
        """
        Delete a work package relation.

        Args:
            relation_id: The relation ID

        Returns:
            bool: True if successful
        """
        await self._request("DELETE", f"/relations/{relation_id}")
        return True

    async def get_work_package_relation(self, relation_id: int) -> Dict:
        """
        Retrieve a specific work package relation by ID.

        Args:
            relation_id: The relation ID

        Returns:
            Dict: Relation data
        """
        return await self._request("GET", f"/relations/{relation_id}")


class OpenProjectMCPServer:
    """MCP Server for OpenProject integration"""

    def __init__(self):
        self.server = Server("openproject-mcp")
        self.client: Optional[OpenProjectClient] = None
        self._setup_handlers()

    def _setup_handlers(self):
        """Register all MCP handlers"""

        @self.server.list_tools()
        async def list_tools() -> List[Tool]:
            """List available tools"""
            return [
                Tool(
                    name="test_connection",
                    description="Test the connection to the OpenProject API",
                    inputSchema={"type": "object", "properties": {}},
                ),
                Tool(
                    name="list_projects",
                    description="List all OpenProject projects",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "active_only": {
                                "type": "boolean",
                                "description": "Show only active projects",
                                "default": True,
                            }
                        },
                    },
                ),
                Tool(
                    name="list_work_packages",
                    description="List work packages with optional pagination",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "project_id": {
                                "type": "integer",
                                "description": "Project ID (optional, for project-specific work packages)",
                            },
                            "status": {
                                "type": "string",
                                "description": "Status filter (open, closed, all)",
                                "enum": ["open", "closed", "all"],
                                "default": "open",
                            },
                            "offset": {
                                "type": "integer",
                                "description": "Starting index for pagination (optional, default: 1)",
                            },
                            "page_size": {
                                "type": "integer",
                                "description": "Number of results per page (optional, max: 100)",
                            },
                        },
                    },
                ),
                Tool(
                    name="list_types",
                    description="List available work package types",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "project_id": {
                                "type": "integer",
                                "description": "Project ID (optional, for project-specific types)",
                            }
                        },
                    },
                ),
                Tool(
                    name="create_work_package",
                    description="Create a new work package with optional date fields",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "project_id": {
                                "type": "integer",
                                "description": "Project ID",
                            },
                            "subject": {
                                "type": "string",
                                "description": "Work package title",
                            },
                            "description": {
                                "type": "string",
                                "description": "Description (Markdown supported)",
                            },
                            "type_id": {
                                "type": "integer",
                                "description": "Type ID (e.g., 1 for Task, 2 for Bug)",
                            },
                            "priority_id": {
                                "type": "integer",
                                "description": "Priority ID (optional)",
                            },
                            "assignee_id": {
                                "type": "integer",
                                "description": "Assignee user ID (optional)",
                            },
                            "start_date": {
                                "type": "string",
                                "description": "Start date in ISO 8601 format: YYYY-MM-DD (optional)",
                            },
                            "due_date": {
                                "type": "string",
                                "description": "Due date in ISO 8601 format: YYYY-MM-DD (optional)",
                            },
                            "date": {
                                "type": "string",
                                "description": "Date for milestones in ISO 8601 format: YYYY-MM-DD (optional)",
                            },
                        },
                        "required": ["project_id", "subject", "type_id"],
                    },
                ),
                Tool(
                    name="list_users",
                    description="List all users",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "active_only": {
                                "type": "boolean",
                                "description": "Show only active users",
                                "default": True,
                            }
                        },
                    },
                ),
                Tool(
                    name="get_user",
                    description="Get detailed information about a specific user",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "user_id": {"type": "integer", "description": "User ID"}
                        },
                        "required": ["user_id"],
                    },
                ),
                Tool(
                    name="list_memberships",
                    description="List project memberships",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "project_id": {
                                "type": "integer",
                                "description": "Project ID (optional, for project-specific memberships)",
                            },
                            "user_id": {
                                "type": "integer",
                                "description": "User ID (optional, for user-specific memberships)",
                            },
                        },
                    },
                ),
                Tool(
                    name="list_statuses",
                    description="List available work package statuses",
                    inputSchema={"type": "object", "properties": {}},
                ),
                Tool(
                    name="list_priorities",
                    description="List available work package priorities",
                    inputSchema={"type": "object", "properties": {}},
                ),
                Tool(
                    name="get_work_package",
                    description="Get detailed information about a specific work package",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "work_package_id": {
                                "type": "integer",
                                "description": "Work package ID",
                            }
                        },
                        "required": ["work_package_id"],
                    },
                ),
                Tool(
                    name="update_work_package",
                    description="Update an existing work package including dates",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "work_package_id": {
                                "type": "integer",
                                "description": "Work package ID",
                            },
                            "subject": {
                                "type": "string",
                                "description": "Work package title (optional)",
                            },
                            "description": {
                                "type": "string",
                                "description": "Description (Markdown supported, optional)",
                            },
                            "type_id": {
                                "type": "integer",
                                "description": "Type ID (optional)",
                            },
                            "status_id": {
                                "type": "integer",
                                "description": "Status ID (optional)",
                            },
                            "priority_id": {
                                "type": "integer",
                                "description": "Priority ID (optional)",
                            },
                            "assignee_id": {
                                "type": "integer",
                                "description": "Assignee user ID (optional)",
                            },
                            "percentage_done": {
                                "type": "integer",
                                "description": "Completion percentage (0-100, optional)",
                            },
                            "start_date": {
                                "type": "string",
                                "description": "Start date in ISO 8601 format: YYYY-MM-DD (optional)",
                            },
                            "due_date": {
                                "type": "string",
                                "description": "Due date in ISO 8601 format: YYYY-MM-DD (optional)",
                            },
                            "date": {
                                "type": "string",
                                "description": "Date for milestones in ISO 8601 format: YYYY-MM-DD (optional)",
                            },
                        },
                        "required": ["work_package_id"],
                    },
                ),
                Tool(
                    name="delete_work_package",
                    description="Delete a work package",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "work_package_id": {
                                "type": "integer",
                                "description": "Work package ID",
                            }
                        },
                        "required": ["work_package_id"],
                    },
                ),
                Tool(
                    name="list_time_entries",
                    description="List time entries",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "work_package_id": {
                                "type": "integer",
                                "description": "Work package ID (optional, for work package-specific time entries)",
                            },
                            "user_id": {
                                "type": "integer",
                                "description": "User ID (optional, for user-specific time entries)",
                            },
                        },
                    },
                ),
                Tool(
                    name="create_time_entry",
                    description="Create a new time entry",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "work_package_id": {
                                "type": "integer",
                                "description": "Work package ID",
                            },
                            "hours": {
                                "type": "number",
                                "description": "Hours spent (e.g., 2.5)",
                            },
                            "spent_on": {
                                "type": "string",
                                "description": "Date when time was spent (YYYY-MM-DD format)",
                            },
                            "comment": {
                                "type": "string",
                                "description": "Comment/description (optional)",
                            },
                            "activity_id": {
                                "type": "integer",
                                "description": "Activity ID (optional)",
                            },
                        },
                        "required": ["work_package_id", "hours", "spent_on"],
                    },
                ),
                Tool(
                    name="update_time_entry",
                    description="Update an existing time entry",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "time_entry_id": {
                                "type": "integer",
                                "description": "Time entry ID",
                            },
                            "hours": {
                                "type": "number",
                                "description": "Hours spent (e.g., 2.5, optional)",
                            },
                            "spent_on": {
                                "type": "string",
                                "description": "Date when time was spent (YYYY-MM-DD format, optional)",
                            },
                            "comment": {
                                "type": "string",
                                "description": "Comment/description (optional)",
                            },
                            "activity_id": {
                                "type": "integer",
                                "description": "Activity ID (optional)",
                            },
                        },
                        "required": ["time_entry_id"],
                    },
                ),
                Tool(
                    name="delete_time_entry",
                    description="Delete a time entry",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "time_entry_id": {
                                "type": "integer",
                                "description": "Time entry ID",
                            }
                        },
                        "required": ["time_entry_id"],
                    },
                ),
                Tool(
                    name="list_time_entry_activities",
                    description="List available time entry activities",
                    inputSchema={"type": "object", "properties": {}},
                ),
                Tool(
                    name="list_versions",
                    description="List project versions/milestones",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "project_id": {
                                "type": "integer",
                                "description": "Project ID (optional, for project-specific versions)",
                            }
                        },
                    },
                ),
                Tool(
                    name="create_version",
                    description="Create a new project version/milestone",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "project_id": {
                                "type": "integer",
                                "description": "Project ID",
                            },
                            "name": {"type": "string", "description": "Version name"},
                            "description": {
                                "type": "string",
                                "description": "Version description (optional)",
                            },
                            "start_date": {
                                "type": "string",
                                "description": "Start date (YYYY-MM-DD format, optional)",
                            },
                            "end_date": {
                                "type": "string",
                                "description": "End date (YYYY-MM-DD format, optional)",
                            },
                            "status": {
                                "type": "string",
                                "description": "Version status (open, locked, closed, optional)",
                            },
                        },
                        "required": ["project_id", "name"],
                    },
                ),
                Tool(
                    name="check_permissions",
                    description="Check current user permissions and capabilities",
                    inputSchema={"type": "object", "properties": {}},
                ),
                Tool(
                    name="create_project",
                    description="Create a new project",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "name": {"type": "string", "description": "Project name"},
                            "identifier": {
                                "type": "string",
                                "description": "Project identifier (unique)",
                            },
                            "description": {
                                "type": "string",
                                "description": "Project description (optional)",
                            },
                            "public": {
                                "type": "boolean",
                                "description": "Whether the project is public (optional)",
                            },
                            "status": {
                                "type": "string",
                                "description": "Project status (optional)",
                            },
                            "parent_id": {
                                "type": "integer",
                                "description": "Parent project ID (optional)",
                            },
                        },
                        "required": ["name", "identifier"],
                    },
                ),
                Tool(
                    name="update_project",
                    description="Update an existing project",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "project_id": {
                                "type": "integer",
                                "description": "Project ID",
                            },
                            "name": {
                                "type": "string",
                                "description": "Project name (optional)",
                            },
                            "identifier": {
                                "type": "string",
                                "description": "Project identifier (optional)",
                            },
                            "description": {
                                "type": "string",
                                "description": "Project description (optional)",
                            },
                            "public": {
                                "type": "boolean",
                                "description": "Whether the project is public (optional)",
                            },
                            "status": {
                                "type": "string",
                                "description": "Project status (optional)",
                            },
                            "parent_id": {
                                "type": "integer",
                                "description": "Parent project ID (optional)",
                            },
                        },
                        "required": ["project_id"],
                    },
                ),
                Tool(
                    name="delete_project",
                    description="Delete a project",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "project_id": {
                                "type": "integer",
                                "description": "Project ID",
                            }
                        },
                        "required": ["project_id"],
                    },
                ),
                Tool(
                    name="get_project",
                    description="Get detailed information about a specific project",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "project_id": {
                                "type": "integer",
                                "description": "Project ID",
                            }
                        },
                        "required": ["project_id"],
                    },
                ),
                Tool(
                    name="create_membership",
                    description="Create a new project membership",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "project_id": {
                                "type": "integer",
                                "description": "Project ID",
                            },
                            "user_id": {
                                "type": "integer",
                                "description": "User ID (required if group_id not provided)",
                            },
                            "group_id": {
                                "type": "integer",
                                "description": "Group ID (required if user_id not provided)",
                            },
                            "role_ids": {
                                "type": "array",
                                "items": {"type": "integer"},
                                "description": "Array of role IDs",
                            },
                            "role_id": {
                                "type": "integer",
                                "description": "Single role ID (alternative to role_ids)",
                            },
                            "notification_message": {
                                "type": "string",
                                "description": "Optional notification message",
                            },
                        },
                        "required": ["project_id"],
                    },
                ),
                Tool(
                    name="update_membership",
                    description="Update an existing membership",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "membership_id": {
                                "type": "integer",
                                "description": "Membership ID",
                            },
                            "role_ids": {
                                "type": "array",
                                "items": {"type": "integer"},
                                "description": "Array of role IDs",
                            },
                            "role_id": {
                                "type": "integer",
                                "description": "Single role ID (alternative to role_ids)",
                            },
                            "notification_message": {
                                "type": "string",
                                "description": "Optional notification message",
                            },
                        },
                        "required": ["membership_id"],
                    },
                ),
                Tool(
                    name="delete_membership",
                    description="Delete a membership",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "membership_id": {
                                "type": "integer",
                                "description": "Membership ID",
                            }
                        },
                        "required": ["membership_id"],
                    },
                ),
                Tool(
                    name="get_membership",
                    description="Get detailed information about a specific membership",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "membership_id": {
                                "type": "integer",
                                "description": "Membership ID",
                            }
                        },
                        "required": ["membership_id"],
                    },
                ),
                Tool(
                    name="list_project_members",
                    description="List all members of a specific project",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "project_id": {
                                "type": "integer",
                                "description": "Project ID",
                            }
                        },
                        "required": ["project_id"],
                    },
                ),
                Tool(
                    name="list_user_projects",
                    description="List all projects a specific user is assigned to",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "user_id": {"type": "integer", "description": "User ID"}
                        },
                        "required": ["user_id"],
                    },
                ),
                Tool(
                    name="list_roles",
                    description="List all available roles",
                    inputSchema={"type": "object", "properties": {}},
                ),
                Tool(
                    name="get_role",
                    description="Get detailed information about a specific role",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "role_id": {"type": "integer", "description": "Role ID"}
                        },
                        "required": ["role_id"],
                    },
                ),
                Tool(
                    name="set_work_package_parent",
                    description="Set a parent for a work package (create parent-child relationship)",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "work_package_id": {
                                "type": "integer",
                                "description": "Work package ID to become a child",
                            },
                            "parent_id": {
                                "type": "integer",
                                "description": "Work package ID to become the parent",
                            },
                        },
                        "required": ["work_package_id", "parent_id"],
                    },
                ),
                Tool(
                    name="remove_work_package_parent",
                    description="Remove parent relationship from a work package (make it top-level)",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "work_package_id": {
                                "type": "integer",
                                "description": "Work package ID to remove parent from",
                            }
                        },
                        "required": ["work_package_id"],
                    },
                ),
                Tool(
                    name="list_work_package_children",
                    description="List all child work packages of a parent",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "parent_id": {
                                "type": "integer",
                                "description": "Parent work package ID",
                            },
                            "include_descendants": {
                                "type": "boolean",
                                "description": "Include grandchildren and all descendants (default: false)",
                                "default": False,
                            },
                        },
                        "required": ["parent_id"],
                    },
                ),
                Tool(
                    name="create_work_package_relation",
                    description="Create a relationship between work packages",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "from_id": {
                                "type": "integer",
                                "description": "Source work package ID",
                            },
                            "to_id": {
                                "type": "integer",
                                "description": "Target work package ID",
                            },
                            "relation_type": {
                                "type": "string",
                                "description": "Relation type",
                                "enum": [
                                    "blocks",
                                    "follows",
                                    "precedes",
                                    "relates",
                                    "duplicates",
                                    "includes",
                                    "requires",
                                    "partof",
                                ],
                            },
                            "lag": {
                                "type": "integer",
                                "description": "Lag in working days (optional, for follows/precedes)",
                            },
                            "description": {
                                "type": "string",
                                "description": "Optional description of the relation",
                            },
                        },
                        "required": ["from_id", "to_id", "relation_type"],
                    },
                ),
                Tool(
                    name="list_work_package_relations",
                    description="List work package relations with optional filtering",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "work_package_id": {
                                "type": "integer",
                                "description": "Filter relations involving this work package ID (optional)",
                            },
                            "relation_type": {
                                "type": "string",
                                "description": "Filter by relation type (optional)",
                                "enum": [
                                    "blocks",
                                    "follows",
                                    "precedes",
                                    "relates",
                                    "duplicates",
                                    "includes",
                                    "requires",
                                    "partof",
                                ],
                            },
                        },
                    },
                ),
                Tool(
                    name="update_work_package_relation",
                    description="Update an existing work package relation",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "relation_id": {
                                "type": "integer",
                                "description": "Relation ID",
                            },
                            "relation_type": {
                                "type": "string",
                                "description": "New relation type (optional)",
                                "enum": [
                                    "blocks",
                                    "follows",
                                    "precedes",
                                    "relates",
                                    "duplicates",
                                    "includes",
                                    "requires",
                                    "partof",
                                ],
                            },
                            "lag": {
                                "type": "integer",
                                "description": "Lag in working days (optional, for follows/precedes)",
                            },
                            "description": {
                                "type": "string",
                                "description": "Optional description of the relation",
                            },
                        },
                        "required": ["relation_id"],
                    },
                ),
                Tool(
                    name="delete_work_package_relation",
                    description="Delete a work package relation",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "relation_id": {
                                "type": "integer",
                                "description": "Relation ID",
                            }
                        },
                        "required": ["relation_id"],
                    },
                ),
                Tool(
                    name="get_work_package_relation",
                    description="Get detailed information about a specific work package relation",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "relation_id": {
                                "type": "integer",
                                "description": "Relation ID",
                            }
                        },
                        "required": ["relation_id"],
                    },
                ),
                Tool(
                    name="create_comment",
                    description="Create a comment on a work package",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "work_package_id": {
                                "type": "integer",
                                "description": "Work package ID",
                            },
                            "comment": {
                                "type": "string",
                                "description": "Comment text",
                            },
                        },
                        "required": ["work_package_id", "comment"],
                    },
                ),
                Tool(
                    name="get_comment",
                    description="Get comments (comments and changes) for a work package",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "work_package_id": {
                                "type": "integer",
                                "description": "Work package ID",
                            },
                        },
                        "required": ["work_package_id"],
                    },
                ),
                Tool(
                    name="update_comment",
                    description="Update a comment",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "comment_id": {
                                "type": "integer",
                                "description": "Comment ID",
                            },
                            "comment": {
                                "type": "string",
                                "description": "New comment text",
                            },
                        },
                        "required": ["comment_id", "comment"],
                    },
                ),
                Tool(
                    name="export_view",
                    description="Export a user-defined view (query) to XLSX format with columns and filters exactly as shown in web",
                    inputSchema={
                        "type": "object",
                        "properties": {
                            "query_id": {
                                "type": "integer",
                                "description": "Query/View ID to export (optional if query_name provided)",
                            },
                            "query_name": {
                                "type": "string",
                                "description": "Query/View name to export (optional if query_id provided)",
                            },
                            "project_id": {
                                "type": "integer",
                                "description": "Project ID to search for view (optional if project_name provided)",
                            },
                            "project_name": {
                                "type": "string",
                                "description": "Project name to search for view (optional if project_id provided)",
                            },
                            "output_file": {
                                "type": "string",
                                "description": "Output file path (e.g., /tmp/export.xlsx)",
                            },
                        },
                        "required": ["output_file"],
                    },
                ),
            ]

        @self.server.call_tool()
        async def call_tool(name: str, arguments: Dict[str, Any]) -> List[TextContent]:
            """Execute a tool"""
            if not self.client:
                return [
                    TextContent(
                        type="text",
                        text="Error: OpenProject Client not initialized. Please set environment variables:\n"
                        "- OPENPROJECT_URL=https://your-instance.openproject.com\n"
                        "- OPENPROJECT_API_KEY=your-api-key",
                    )
                ]

            try:
                if name == "test_connection":
                    result = await self.client.test_connection()

                    text = "✅ API connection successful!\n\n"
                    if self.client.proxy:
                        text += f"Connected via proxy: {self.client.proxy}\n"
                    text += f"API Version: {result.get('_type', 'Unknown')}\n"
                    text += f"Instance Version: {result.get('instanceVersion', 'Unknown')}\n"

                    return [TextContent(type="text", text=text)]

                elif name == "list_projects":
                    filters = None
                    if arguments.get("active_only", True):
                        filters = json.dumps(
                            [{"active": {"operator": "=", "values": ["t"]}}]
                        )

                    result = await self.client.get_projects(filters)
                    projects = result.get("_embedded", {}).get("elements", [])

                    if not projects:
                        text = "No projects found."
                    else:
                        text = f"Found {len(projects)} project(s):\n\n"
                        for project in projects:
                            text += f"- **{project['name']}** (ID: {project['id']})\n"
                            if project.get("description", {}).get("raw"):
                                text += f"  {project['description']['raw']}\n"
                            text += f"  Status: {'Active' if project.get('active') else 'Inactive'}\n"
                            text += f"  Public: {'Yes' if project.get('public') else 'No'}\n\n"

                    return [TextContent(type="text", text=text)]

                elif name == "list_work_packages":
                    project_id = arguments.get("project_id")
                    status = arguments.get("status", "open")
                    offset = arguments.get("offset")
                    page_size = arguments.get("page_size")

                    filters = None
                    if status == "open":
                        filters = json.dumps(
                            [{"status_id": {"operator": "o", "values": None}}]
                        )
                    elif status == "closed":
                        filters = json.dumps(
                            [{"status_id": {"operator": "c", "values": None}}]
                        )

                    result = await self.client.get_work_packages(
                        project_id, filters, offset, page_size
                    )
                    work_packages = result.get("_embedded", {}).get("elements", [])

                    # Get pagination info from result
                    total = result.get("total", len(work_packages))
                    count = result.get("count", len(work_packages))
                    page_size_actual = result.get("pageSize", page_size or 20)
                    offset_actual = result.get("offset", offset or 1)

                    if not work_packages:
                        text = "No work packages found."
                    else:
                        # Show pagination info
                        text = f"Found {total} work package(s) (showing {count} results"
                        if offset or page_size:
                            text += f", offset: {offset_actual}, pageSize: {page_size_actual}"
                        text += "):\n\n"

                        for wp in work_packages:
                            text += f"- **{wp.get('subject', 'No title')}** (#{wp.get('id', 'N/A')})\n"

                            if "_embedded" in wp:
                                embedded = wp["_embedded"]
                                if "type" in embedded:
                                    text += f"  Type: {embedded['type'].get('name', 'Unknown')}\n"
                                if "status" in embedded:
                                    text += f"  Status: {embedded['status'].get('name', 'Unknown')}\n"
                                if "project" in embedded:
                                    text += f"  Project: {embedded['project'].get('name', 'Unknown')}\n"
                                if "assignee" in embedded and embedded["assignee"]:
                                    text += f"  Assignee: {embedded['assignee'].get('name', 'Unassigned')}\n"

                            if "percentageDone" in wp:
                                text += f"  Progress: {wp['percentageDone']}%\n"

                            text += "\n"

                    return [TextContent(type="text", text=text)]

                elif name == "list_types":
                    result = await self.client.get_types(arguments.get("project_id"))
                    types = result.get("_embedded", {}).get("elements", [])

                    if not types:
                        text = "No work package types found."
                    else:
                        text = "Available work package types:\n\n"
                        for type_item in types:
                            text += f"- **{type_item.get('name', 'Unnamed')}** (ID: {type_item.get('id', 'N/A')})\n"
                            if type_item.get("isDefault"):
                                text += "  ✓ Default type\n"
                            if type_item.get("isMilestone"):
                                text += "  ✓ Milestone\n"
                            text += "\n"

                    return [TextContent(type="text", text=text)]

                elif name == "create_work_package":
                    try:
                        data = {
                            "project": arguments["project_id"],
                            "subject": arguments["subject"],
                            "type": arguments["type_id"],
                        }

                        # Add optional fields
                        for field in ["description", "priority_id", "assignee_id"]:
                            if field in arguments:
                                data[field] = arguments[field]

                        # Add date fields (map from snake_case to camelCase)
                        if "start_date" in arguments:
                            data["startDate"] = arguments["start_date"]
                        if "due_date" in arguments:
                            data["dueDate"] = arguments["due_date"]
                        if "date" in arguments:
                            data["date"] = arguments["date"]

                        result = await self.client.create_work_package(data)

                        text = f"✅ Work package created successfully:\n\n"
                        text += f"- **Title**: {result.get('subject', 'N/A')}\n"
                        text += f"- **ID**: #{result.get('id', 'N/A')}\n"

                        if "_embedded" in result:
                            embedded = result["_embedded"]
                            if "type" in embedded:
                                text += f"- **Type**: {embedded['type'].get('name', 'Unknown')}\n"
                            if "status" in embedded:
                                text += f"- **Status**: {embedded['status'].get('name', 'Unknown')}\n"
                            if "project" in embedded:
                                text += f"- **Project**: {embedded['project'].get('name', 'Unknown')}\n"

                        return [TextContent(type="text", text=text)]

                    except Exception as e:
                        error_msg = str(e)
                        if "403" in error_msg:
                            # Check user permissions for better error message
                            try:
                                user_info = await self.client.check_permissions()
                                text = f"❌ Permission Error: Cannot create work packages.\n\n"
                                text += f"**Current User**: {user_info.get('name', 'Unknown')}\n"
                                text += f"**Admin**: {'Yes' if user_info.get('admin') else 'No'}\n\n"
                                text += "**Possible Solutions:**\n"
                                text += "1. Contact your OpenProject administrator to grant work package creation permissions\n"
                                text += "2. Ensure you have 'Create work packages' permission in the target project\n"
                                text += "3. Check if the project allows work package creation\n"
                                text += f"4. Verify project ID {arguments['project_id']} exists and you have access\n\n"
                                text += f"**Technical Error**: {error_msg}"
                                return [TextContent(type="text", text=text)]
                            except:
                                pass

                        # Default error handling
                        text = f"❌ Failed to create work package: {error_msg}"
                        return [TextContent(type="text", text=text)]

                elif name == "create_comment":
                    work_package_id = arguments["work_package_id"]
                    comment = arguments["comment"]
                    
                    result = await self.client.create_comment(work_package_id, comment)
                    
                    text = f"✅ Comment created successfully on work package #{work_package_id}\n\n"
                    text += f"- **Comment**: {comment}\n"
                    if "id" in result:
                        text += f"- **Comment ID**: {result.get('id', 'N/A')}\n"
                    if "createdAt" in result:
                        text += f"- **Created**: {result.get('createdAt', 'N/A')}\n"
                    
                    return [TextContent(type="text", text=text)]

                elif name == "get_comment":
                    work_package_id = arguments["work_package_id"]
                    result = await self.client.get_comment(work_package_id)
                    
                    comments = result.get("_embedded", {}).get("elements", [])
                    
                    if not comments:
                        text = f"No comments found for work package #{work_package_id}."
                    else:
                        text = f"Found {len(comments)} comment(s) for work package #{work_package_id}:\n\n"
                        for comment_item in comments:
                            text += f"**Comment #{comment_item.get('id', 'N/A')}**\n"
                            text += f"- **Version**: {comment_item.get('version', 'N/A')}\n"
                            text += f"- **Created**: {comment_item.get('createdAt', 'N/A')}\n"
                            
                            # User information
                            if "_links" in comment_item and "user" in comment_item["_links"]:
                                user_link = comment_item["_links"]["user"]
                                if "title" in user_link:
                                    text += f"- **User**: {user_link['title']}\n"
                            
                            # Comment
                            if "comment" in comment_item and comment_item["comment"]:
                                comment_text = comment_item["comment"].get("raw", "")
                                if comment_text:
                                    text += f"- **Comment**: {comment_text}\n"
                            
                            # Details about changes
                            if "details" in comment_item and comment_item["details"]:
                                details = comment_item["details"]
                                if details:
                                    text += f"- **Changes**: {len(details)} field(s) modified\n"
                            
                            text += "\n"
                    
                    return [TextContent(type="text", text=text)]

                elif name == "update_comment":
                    comment_id = arguments["comment_id"]
                    comment = arguments["comment"]
                    
                    result = await self.client.update_comment(comment_id, comment)
                    
                    text = f"✅ Comment #{comment_id} updated successfully\n\n"
                    text += f"- **New Comment**: {comment}\n"
                    if "updatedAt" in result:
                        text += f"- **Updated**: {result.get('updatedAt', 'N/A')}\n"
                    
                    return [TextContent(type="text", text=text)]

                elif name == "export_view":
                    if not OPENPYXL_AVAILABLE:
                        return [TextContent(
                            type="text",
                            text="❌ Error: openpyxl library not installed. Install with: pip install openpyxl"
                        )]
                    
                    query_id = arguments.get("query_id")
                    query_name = arguments.get("query_name")
                    project_id = arguments.get("project_id")
                    project_name = arguments.get("project_name")
                    output_file = arguments["output_file"]
                    
                    # Validate that at least one identifier is provided
                    if not query_id and not query_name:
                        return [TextContent(
                            type="text",
                            text="❌ Error: Please provide either query_id or query_name"
                        )]
                    
                    try:
                        # If project_name is provided, search for the project
                        if project_name and not project_id:
                            projects_result = await self.client.get_projects()
                            projects = projects_result.get("_embedded", {}).get("elements", [])
                            
                            # Normalize function for robust matching
                            import re
                            def normalize_name(s: str) -> str:
                                """Normalize name: lowercase, strip, collapse whitespace, remove special chars"""
                                s = s.lower().strip()
                                s = re.sub(r'\s+', ' ', s)  # Collapse whitespace
                                s = re.sub(r'[^\w\s-]', '', s)  # Remove special chars except dash
                                return s
                            
                            target_normalized = normalize_name(project_name)
                            
                            # 1. Try exact normalized match
                            matching_project = None
                            for p in projects:
                                if normalize_name(p.get("name", "")) == target_normalized:
                                    matching_project = p
                                    break
                            
                            # 2. Try partial normalized match (contains)
                            if not matching_project:
                                for p in projects:
                                    if target_normalized in normalize_name(p.get("name", "")):
                                        matching_project = p
                                        break
                            
                            # 3. Try reverse partial match (project name contains search term)
                            if not matching_project:
                                for p in projects:
                                    project_normalized = normalize_name(p.get("name", ""))
                                    if project_normalized in target_normalized:
                                        matching_project = p
                                        break
                            
                            if not matching_project:
                                available = ", ".join([f"'{p.get('name')}' (ID: {p.get('id')})" for p in projects[:10]])
                                return [TextContent(
                                    type="text",
                                    text=f"❌ Error: Project '{project_name}' not found.\n\nAvailable projects: {available}"
                                )]
                            
                            project_id = matching_project.get("id")
                        
                        # If query_name is provided, search for the query
                        if query_name and not query_id:
                            try:
                                queries_result = await self.client.get_queries(project_id)
                                queries = queries_result.get("_embedded", {}).get("elements", [])
                                
                                if not queries:
                                    project_info = f" in project ID {project_id}" if project_id else ""
                                    return [TextContent(
                                        type="text",
                                        text=f"❌ Error: No views/queries found{project_info}. The project might not have any saved views."
                                    )]
                                
                                # Search for matching query by name (case-insensitive)
                                matching_query = None
                                for q in queries:
                                    if q.get("name", "").lower() == query_name.lower():
                                        matching_query = q
                                        break
                                
                                if not matching_query:
                                    # Try partial match
                                    for q in queries:
                                        if query_name.lower() in q.get("name", "").lower():
                                            matching_query = q
                                            break
                                
                                if not matching_query:
                                    available = ", ".join([f"'{q.get('name')}' (ID: {q.get('id')})" for q in queries[:15]])
                                    project_info = f" in project ID {project_id}" if project_id else ""
                                    return [TextContent(
                                        type="text",
                                        text=f"❌ Error: Query/View '{query_name}' not found{project_info}.\n\nAvailable views ({len(queries)} total): {available}"
                                    )]
                                
                                query_id = matching_query.get("id")
                            
                            except Exception as e:
                                error_msg = str(e)
                                return [TextContent(
                                    type="text",
                                    text=f"❌ Error searching for queries: {error_msg}"
                                )]
                        
                        # Get query details
                        try:
                            query = await self.client.get_query(query_id)
                            query_name = query.get("name", f"Query {query_id}")
                        except Exception as e:
                            return [TextContent(
                                type="text",
                                text=f"❌ Error getting query details for ID {query_id}: {str(e)}\n\nThe query might be a default system query that doesn't support direct access. Try using a custom saved view instead."
                            )]
                        
                        # Get query results (work packages)
                        try:
                            results = await self.client.get_query_results(query_id)
                            work_packages = results.get("_embedded", {}).get("elements", [])
                        except Exception as e:
                            return [TextContent(
                                type="text",
                                text=f"❌ Error getting query results for '{query_name}' (ID {query_id}): {str(e)}"
                            )]
                        
                        # Get column configuration from query
                        columns = query.get("columns", [])
                        
                        # Create workbook
                        wb = Workbook()
                        ws = wb.active
                        ws.title = query_name[:31]  # Excel sheet name limit
                        
                        # Define column mappings (common columns)
                        column_mapping = {
                            "id": "ID",
                            "subject": "Subject",
                            "type": "Type",
                            "status": "Status",
                            "priority": "Priority",
                            "assignee": "Assignee",
                            "author": "Author",
                            "startDate": "Start Date",
                            "dueDate": "Finish Date",
                            "estimatedTime": "Work",
                            "spentTime": "Spent",
                            "percentageDone": "Progress (%)",
                            "createdAt": "Created",
                            "updatedAt": "Updated",
                            "project": "Project",
                            "description": "Description",
                            "budget": "Budget",
                            "costObject": "Budget",
                        }
                        
                        # Extract column names
                        if columns:
                            column_ids = [col.get("id") or col.get("_links", {}).get("self", {}).get("href", "").split("/")[-1] for col in columns]
                        else:
                            # Default columns if not specified
                            column_ids = ["id", "subject", "type", "status", "assignee", "dueDate"]
                        
                        # Ensure essential columns are always included
                        essential_columns = ["startDate", "dueDate", "estimatedTime", "spentTime", "budget"]
                        for essential_col in essential_columns:
                            if essential_col not in column_ids:
                                column_ids.append(essential_col)
                        
                        # Write header row
                        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        header_font = Font(bold=True, color="FFFFFF")
                        
                        for col_idx, col_id in enumerate(column_ids, start=1):
                            cell = ws.cell(row=1, column=col_idx)
                            cell.value = column_mapping.get(col_id, col_id.title())
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                        
                        # Build hierarchy map and reorder work packages
                        parent_map = {}  # wp_id -> parent_id
                        children_map = {}  # parent_id -> [child_ids]
                        wp_map = {}  # wp_id -> work package object
                        
                        for wp in work_packages:
                            wp_id = wp.get("id")
                            wp_map[wp_id] = wp
                            
                            parent_link = wp.get("_links", {}).get("parent", {}).get("href", "")
                            if parent_link:
                                parent_id = int(parent_link.split("/")[-1])
                                parent_map[wp_id] = parent_id
                                
                                if parent_id not in children_map:
                                    children_map[parent_id] = []
                                children_map[parent_id].append(wp_id)
                        
                        # Function to recursively build ordered list with hierarchy
                        def build_ordered_hierarchy(wp_id, depth=0):
                            """Recursively build list of (work_package, depth) tuples"""
                            result = []
                            if wp_id in wp_map:
                                result.append((wp_map[wp_id], depth))
                                # Add children recursively
                                if wp_id in children_map:
                                    for child_id in children_map[wp_id]:
                                        result.extend(build_ordered_hierarchy(child_id, depth + 1))
                            return result
                        
                        # Find root work packages (those without parents in the current result set)
                        root_ids = []
                        for wp_id in wp_map.keys():
                            if wp_id not in parent_map or parent_map[wp_id] not in wp_map:
                                # This is a root (no parent) or parent is outside this result set
                                root_ids.append(wp_id)
                        
                        # Build ordered list maintaining hierarchy
                        ordered_work_packages = []
                        for root_id in root_ids:
                            ordered_work_packages.extend(build_ordered_hierarchy(root_id, 0))
                        
                        # Write data rows with hierarchy support
                        for row_idx, (wp, indent_level) in enumerate(ordered_work_packages, start=2):
                            for col_idx, col_id in enumerate(column_ids, start=1):
                                value = ""
                                
                                # Extract value based on column type
                                if col_id == "id":
                                    value = wp.get("id", "")
                                elif col_id == "subject":
                                    # Add indentation to subject for hierarchy visualization
                                    subject = wp.get("subject", "")
                                    if indent_level > 0:
                                        value = "  " * indent_level + "↳ " + subject
                                    else:
                                        value = subject
                                elif col_id == "description":
                                    desc = wp.get("description", {})
                                    if isinstance(desc, dict):
                                        value = desc.get("raw", "")
                                    else:
                                        value = desc or ""
                                elif col_id in ["startDate", "dueDate", "date"]:
                                    value = wp.get(col_id, "")
                                elif col_id == "percentageDone":
                                    value = wp.get("percentageDone", 0)
                                elif col_id in ["type", "status", "priority", "assignee", "author", "project"]:
                                    # Extract from embedded or links
                                    embedded = wp.get("_embedded", {})
                                    if col_id in embedded:
                                        value = embedded[col_id].get("name", "") or embedded[col_id].get("title", "")
                                    else:
                                        # Try from _links
                                        links = wp.get("_links", {})
                                        if col_id in links:
                                            value = links[col_id].get("title", "")
                                elif col_id in ["createdAt", "updatedAt"]:
                                    value = wp.get(col_id, "")
                                elif col_id in ["estimatedTime", "spentTime"]:
                                    time_val = wp.get(col_id, "")
                                    if time_val and "PT" in str(time_val):
                                        # Parse ISO duration PT8H to 8h
                                        value = time_val.replace("PT", "").replace("H", "h")
                                    else:
                                        value = time_val or ""
                                elif col_id == "budget":
                                    # Budget might be in _links or _embedded
                                    embedded = wp.get("_embedded", {})
                                    links = wp.get("_links", {})
                                    
                                    # Check various possible locations for budget
                                    if "budget" in embedded:
                                        value = embedded["budget"].get("name", "") or embedded["budget"].get("title", "")
                                    elif "budget" in links:
                                        value = links["budget"].get("title", "")
                                    elif "costObject" in embedded:
                                        # Sometimes called costObject
                                        value = embedded["costObject"].get("name", "") or embedded["costObject"].get("subject", "")
                                    elif "costObject" in links:
                                        value = links["costObject"].get("title", "")
                                    else:
                                        # Direct field access
                                        value = wp.get("budget", "") or wp.get("costObject", "")
                                elif col_id == "costObject":
                                    # Explicit cost object handling
                                    embedded = wp.get("_embedded", {})
                                    links = wp.get("_links", {})
                                    if "costObject" in embedded:
                                        value = embedded["costObject"].get("name", "") or embedded["costObject"].get("subject", "")
                                    elif "costObject" in links:
                                        value = links["costObject"].get("title", "")
                                    else:
                                        value = wp.get("costObject", "")
                                else:
                                    # Generic field access
                                    value = wp.get(col_id, "")
                                
                                cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
                                
                                # Apply indentation styling for subject column
                                if col_id == "subject" and indent_level > 0:
                                    cell.font = Font(italic=True)
                                    # Light gray background for child items
                                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                        
                        # Auto-adjust column widths
                        for col_idx, col_id in enumerate(column_ids, start=1):
                            max_length = len(column_mapping.get(col_id, col_id.title()))
                            for row_idx in range(2, len(ordered_work_packages) + 2):
                                cell_value = ws.cell(row=row_idx, column=col_idx).value
                                if cell_value:
                                    max_length = max(max_length, len(str(cell_value)[:50]))  # Limit to 50 chars
                            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)
                        
                        # Save workbook
                        wb.save(output_file)
                        
                        text = f"✅ View exported successfully\n\n"
                        text += f"- **Query**: {query_name}\n"
                        text += f"- **Work Packages**: {len(ordered_work_packages)}\n"
                        text += f"- **Columns**: {len(column_ids)}\n"
                        text += f"- **Output File**: {output_file}\n"
                        
                        return [TextContent(type="text", text=text)]
                    
                    except Exception as e:
                        return [TextContent(
                            type="text",
                            text=f"❌ Error exporting view: {str(e)}"
                        )]

                elif name == "list_users":
                    filters = None
                    if arguments.get("active_only", True):
                        filters = json.dumps(
                            [{"status": {"operator": "=", "values": ["active"]}}]
                        )

                    result = await self.client.get_users(filters)
                    users = result.get("_embedded", {}).get("elements", [])

                    if not users:
                        text = "No users found."
                    else:
                        text = f"Found {len(users)} user(s):\n\n"
                        for user in users:
                            text += f"- **{user.get('name', 'Unnamed')}** (ID: {user.get('id', 'N/A')})\n"
                            text += f"  Email: {user.get('email', 'N/A')}\n"
                            text += f"  Status: {user.get('status', 'Unknown')}\n"
                            if user.get("admin"):
                                text += "  ✓ Administrator\n"
                            text += "\n"

                    return [TextContent(type="text", text=text)]

                elif name == "get_user":
                    user_id = arguments["user_id"]
                    result = await self.client.get_user(user_id)

                    text = f"**User Details:**\n\n"
                    text += f"- **Name**: {result.get('name', 'N/A')}\n"
                    text += f"- **ID**: {result.get('id', 'N/A')}\n"
                    text += f"- **Email**: {result.get('email', 'N/A')}\n"
                    text += f"- **Status**: {result.get('status', 'Unknown')}\n"
                    text += f"- **Language**: {result.get('language', 'N/A')}\n"
                    text += f"- **Admin**: {'Yes' if result.get('admin') else 'No'}\n"
                    text += f"- **Created**: {result.get('createdAt', 'N/A')}\n"
                    text += f"- **Updated**: {result.get('updatedAt', 'N/A')}\n"

                    return [TextContent(type="text", text=text)]

                elif name == "list_memberships":
                    project_id = arguments.get("project_id")
                    user_id = arguments.get("user_id")

                    try:
                        result = await self.client.get_memberships(project_id, user_id)
                        memberships = result.get("_embedded", {}).get("elements", [])

                        if not memberships:
                            filter_info = []
                            if project_id:
                                filter_info.append(f"project {project_id}")
                            if user_id:
                                filter_info.append(f"user {user_id}")
                            filter_text = (
                                f" for {' and '.join(filter_info)}"
                                if filter_info
                                else ""
                            )
                            text = f"No memberships found{filter_text}."
                        else:
                            filter_info = []
                            if project_id:
                                filter_info.append(f"project {project_id}")
                            if user_id:
                                filter_info.append(f"user {user_id}")
                            filter_text = (
                                f" ({' and '.join(filter_info)})" if filter_info else ""
                            )

                            text = f"Found {len(memberships)} membership(s){filter_text}:\n\n"
                            for membership in memberships:
                                text += f"- **Membership ID**: {membership.get('id', 'N/A')}\n"

                                if "_embedded" in membership:
                                    embedded = membership["_embedded"]
                                    if "user" in embedded:
                                        text += f"  User: {embedded['user'].get('name', 'Unknown')}\n"
                                    if "project" in embedded:
                                        text += f"  Project: {embedded['project'].get('name', 'Unknown')}\n"
                                    if "roles" in embedded:
                                        roles = [
                                            role.get("name", "Unknown")
                                            for role in embedded["roles"]
                                        ]
                                        text += f"  Roles: {', '.join(roles)}\n"
                                text += "\n"

                        return [TextContent(type="text", text=text)]

                    except Exception as e:
                        error_msg = str(e)
                        if user_id and "user_id" in arguments:
                            text = f"⚠️ User ID filtering may not be supported in this OpenProject instance.\n\n"
                            text += f"**Error with user_id={user_id}**: {error_msg}\n\n"
                            text += "**Workaround**: Try using `list_memberships` without user_id filter, then manually filter results.\n\n"
                            text += "**Alternative**: Use `list_users` to get user details, then check individual project memberships."
                        else:
                            text = f"❌ Failed to retrieve memberships: {error_msg}"

                        return [TextContent(type="text", text=text)]

                elif name == "list_statuses":
                    result = await self.client.get_statuses()
                    statuses = result.get("_embedded", {}).get("elements", [])

                    if not statuses:
                        text = "No statuses found."
                    else:
                        text = "Available work package statuses:\n\n"
                        for status in statuses:
                            text += f"- **{status.get('name', 'Unnamed')}** (ID: {status.get('id', 'N/A')})\n"
                            text += f"  Position: {status.get('position', 'N/A')}\n"
                            if status.get("isDefault"):
                                text += "  ✓ Default status\n"
                            if status.get("isClosed"):
                                text += "  ✓ Closed status\n"
                            text += "\n"

                    return [TextContent(type="text", text=text)]

                elif name == "list_priorities":
                    result = await self.client.get_priorities()
                    priorities = result.get("_embedded", {}).get("elements", [])

                    if not priorities:
                        text = "No priorities found."
                    else:
                        text = "Available work package priorities:\n\n"
                        for priority in priorities:
                            text += f"- **{priority.get('name', 'Unnamed')}** (ID: {priority.get('id', 'N/A')})\n"
                            text += f"  Position: {priority.get('position', 'N/A')}\n"
                            if priority.get("isDefault"):
                                text += "  ✓ Default priority\n"
                            if priority.get("isActive"):
                                text += "  ✓ Active\n"
                            text += "\n"

                    return [TextContent(type="text", text=text)]

                elif name == "get_work_package":
                    work_package_id = arguments["work_package_id"]
                    result = await self.client.get_work_package(work_package_id)

                    text = f"**Work Package Details:**\n\n"
                    text += f"- **ID**: #{result.get('id', 'N/A')}\n"
                    text += f"- **Subject**: {result.get('subject', 'N/A')}\n"
                    text += f"- **Progress**: {result.get('percentageDone', 0)}%\n"
                    text += f"- **Created**: {result.get('createdAt', 'N/A')}\n"
                    text += f"- **Updated**: {result.get('updatedAt', 'N/A')}\n"

                    if "_embedded" in result:
                        embedded = result["_embedded"]
                        if "type" in embedded:
                            text += f"- **Type**: {embedded['type'].get('name', 'Unknown')}\n"
                        if "status" in embedded:
                            text += f"- **Status**: {embedded['status'].get('name', 'Unknown')}\n"
                        if "priority" in embedded:
                            text += f"- **Priority**: {embedded['priority'].get('name', 'Unknown')}\n"
                        if "project" in embedded:
                            text += f"- **Project**: {embedded['project'].get('name', 'Unknown')}\n"
                        if "assignee" in embedded and embedded["assignee"]:
                            text += f"- **Assignee**: {embedded['assignee'].get('name', 'Unassigned')}\n"
                        else:
                            text += f"- **Assignee**: Unassigned\n"

                    if result.get("description", {}).get("raw"):
                        text += f"\n**Description:**\n{result['description']['raw']}\n"

                    return [TextContent(type="text", text=text)]

                elif name == "update_work_package":
                    work_package_id = arguments["work_package_id"]

                    # Prepare update data
                    update_data = {}
                    for field in [
                        "subject",
                        "description",
                        "type_id",
                        "status_id",
                        "priority_id",
                        "assignee_id",
                        "percentage_done",
                    ]:
                        if field in arguments:
                            update_data[field] = arguments[field]

                    # Add date fields (map from snake_case to camelCase)
                    if "start_date" in arguments:
                        update_data["startDate"] = arguments["start_date"]
                    if "due_date" in arguments:
                        update_data["dueDate"] = arguments["due_date"]
                    if "date" in arguments:
                        update_data["date"] = arguments["date"]

                    if not update_data:
                        return [
                            TextContent(
                                type="text", text="❌ No fields provided to update."
                            )
                        ]

                    result = await self.client.update_work_package(
                        work_package_id, update_data
                    )

                    text = (
                        f"✅ Work package #{work_package_id} updated successfully:\n\n"
                    )
                    text += f"- **Subject**: {result.get('subject', 'N/A')}\n"
                    text += f"- **Progress**: {result.get('percentageDone', 0)}%\n"

                    if "_embedded" in result:
                        embedded = result["_embedded"]
                        if "type" in embedded:
                            text += f"- **Type**: {embedded['type'].get('name', 'Unknown')}\n"
                        if "status" in embedded:
                            text += f"- **Status**: {embedded['status'].get('name', 'Unknown')}\n"
                        if "priority" in embedded:
                            text += f"- **Priority**: {embedded['priority'].get('name', 'Unknown')}\n"
                        if "assignee" in embedded and embedded["assignee"]:
                            text += f"- **Assignee**: {embedded['assignee'].get('name', 'Unassigned')}\n"

                    return [TextContent(type="text", text=text)]

                elif name == "delete_work_package":
                    work_package_id = arguments["work_package_id"]

                    success = await self.client.delete_work_package(work_package_id)

                    if success:
                        text = (
                            f"✅ Work package #{work_package_id} deleted successfully."
                        )
                    else:
                        text = f"❌ Failed to delete work package #{work_package_id}."

                    return [TextContent(type="text", text=text)]

                elif name == "list_time_entries":
                    filters = []

                    # Only add user filter (work_package filter is not supported by API)
                    if "user_id" in arguments:
                        filters.append(
                            {
                                "user": {
                                    "operator": "=",
                                    "values": [str(arguments["user_id"])],
                                }
                            }
                        )

                    filter_string = json.dumps(filters) if filters else None
                    result = await self.client.get_time_entries(filter_string)
                    time_entries = result.get("_embedded", {}).get("elements", [])

                    # Client-side filtering by work_package_id if provided
                    if "work_package_id" in arguments:
                        work_package_id = arguments["work_package_id"]
                        time_entries = [
                            entry for entry in time_entries
                            if entry.get("_links", {}).get("workPackage", {}).get("href", "").endswith(f"/{work_package_id}")
                        ]

                    if not time_entries:
                        text = "No time entries found."
                    else:
                        text = f"Found {len(time_entries)} time entrie(s):\n\n"
                        for entry in time_entries:
                            # Parse hours from ISO duration format (PT2.5H)
                            hours_str = entry.get("hours", "PT0H")
                            hours = (
                                hours_str.replace("PT", "").replace("H", "")
                                if "PT" in hours_str
                                else "0"
                            )

                            text += f"- **Time Entry #{entry.get('id', 'N/A')}**\n"
                            text += f"  Hours: {hours}\n"
                            text += f"  Date: {entry.get('spentOn', 'N/A')}\n"

                            if "_embedded" in entry:
                                embedded = entry["_embedded"]
                                if "workPackage" in embedded:
                                    text += f"  Work Package: {embedded['workPackage'].get('subject', 'Unknown')}\n"
                                if "user" in embedded:
                                    text += f"  User: {embedded['user'].get('name', 'Unknown')}\n"
                                if "activity" in embedded:
                                    text += f"  Activity: {embedded['activity'].get('name', 'Unknown')}\n"

                            if entry.get("comment", {}).get("raw"):
                                text += f"  Comment: {entry['comment']['raw']}\n"
                            text += "\n"

                    return [TextContent(type="text", text=text)]

                elif name == "create_time_entry":
                    data = {
                        "work_package_id": arguments["work_package_id"],
                        "hours": arguments["hours"],
                        "spent_on": arguments["spent_on"],
                    }

                    # Add optional fields
                    for field in ["comment", "activity_id"]:
                        if field in arguments:
                            data[field] = arguments[field]

                    result = await self.client.create_time_entry(data)

                    # Parse hours from ISO duration format
                    hours_str = result.get("hours", "PT0H")
                    hours = (
                        hours_str.replace("PT", "").replace("H", "")
                        if "PT" in hours_str
                        else "0"
                    )

                    text = f"✅ Time entry created successfully:\n\n"
                    text += f"- **ID**: #{result.get('id', 'N/A')}\n"
                    text += f"- **Hours**: {hours}\n"
                    text += f"- **Date**: {result.get('spentOn', 'N/A')}\n"

                    if "_embedded" in result:
                        embedded = result["_embedded"]
                        if "workPackage" in embedded:
                            text += f"- **Work Package**: {embedded['workPackage'].get('subject', 'Unknown')}\n"
                        if "activity" in embedded:
                            text += f"- **Activity**: {embedded['activity'].get('name', 'Unknown')}\n"

                    return [TextContent(type="text", text=text)]

                elif name == "update_time_entry":
                    time_entry_id = arguments["time_entry_id"]

                    # Prepare update data
                    update_data = {}
                    for field in ["hours", "spent_on", "comment", "activity_id"]:
                        if field in arguments:
                            update_data[field] = arguments[field]

                    if not update_data:
                        return [
                            TextContent(
                                type="text", text="❌ No fields provided to update."
                            )
                        ]

                    result = await self.client.update_time_entry(
                        time_entry_id, update_data
                    )

                    # Parse hours from ISO duration format
                    hours_str = result.get("hours", "PT0H")
                    hours = (
                        hours_str.replace("PT", "").replace("H", "")
                        if "PT" in hours_str
                        else "0"
                    )

                    text = f"✅ Time entry #{time_entry_id} updated successfully:\n\n"
                    text += f"- **Hours**: {hours}\n"
                    text += f"- **Date**: {result.get('spentOn', 'N/A')}\n"

                    if "_embedded" in result:
                        embedded = result["_embedded"]
                        if "activity" in embedded:
                            text += f"- **Activity**: {embedded['activity'].get('name', 'Unknown')}\n"

                    return [TextContent(type="text", text=text)]

                elif name == "delete_time_entry":
                    time_entry_id = arguments["time_entry_id"]

                    success = await self.client.delete_time_entry(time_entry_id)

                    if success:
                        text = f"✅ Time entry #{time_entry_id} deleted successfully."
                    else:
                        text = f"❌ Failed to delete time entry #{time_entry_id}."

                    return [TextContent(type="text", text=text)]

                elif name == "list_time_entry_activities":
                    try:
                        result = await self.client.get_time_entry_activities()
                        activities = result.get("_embedded", {}).get("elements", [])

                        if not activities:
                            text = "No time entry activities found."
                        else:
                            text = "Available time entry activities:\n\n"
                            for activity in activities:
                                text += f"- **{activity.get('name', 'Unnamed')}** (ID: {activity.get('id', 'N/A')})\n"
                                if activity.get("position"):
                                    text += f"  Position: {activity.get('position')}\n"
                                if activity.get("isDefault"):
                                    text += "  ✓ Default activity\n"
                                text += "\n"

                        return [TextContent(type="text", text=text)]

                    except Exception as e:
                        error_msg = str(e)
                        if "404" in error_msg:
                            # Provide fallback with discovered activity IDs
                            text = "⚠️ Time entry activities endpoint not available, but activities can still be used!\n\n"
                            text += (
                                "**Available Time Entry Activities (Discovered):**\n\n"
                            )
                            text += "- **Management** (ID: 1)\n"
                            text += "  Administrative and planning tasks\n\n"
                            text += "- **Specification** (ID: 2)\n"
                            text += "  Requirements and documentation\n\n"
                            text += "- **Development** (ID: 3)\n"
                            text += "  Coding and implementation\n\n"
                            text += "- **Testing** (ID: 4)\n"
                            text += "  Quality assurance and testing\n\n"
                            text += "**Usage**: Use these activity IDs when creating time entries with the `activity_id` parameter.\n\n"
                            text += "**Example**: `create_time_entry` with `activity_id: 3` for Development work\n\n"
                            text += f"**Technical Note**: Endpoint returned 404, but activities are functional: {error_msg}"
                        else:
                            text = f"❌ Failed to retrieve time entry activities: {error_msg}"

                        return [TextContent(type="text", text=text)]

                elif name == "list_versions":
                    project_id = arguments.get("project_id")
                    result = await self.client.get_versions(project_id)
                    versions = result.get("_embedded", {}).get("elements", [])

                    if not versions:
                        text = "No versions found."
                    else:
                        text = f"Found {len(versions)} version(s):\n\n"
                        for version in versions:
                            text += f"- **{version.get('name', 'Unnamed')}** (ID: {version.get('id', 'N/A')})\n"
                            text += f"  Status: {version.get('status', 'Unknown')}\n"

                            if version.get("startDate"):
                                text += f"  Start Date: {version.get('startDate')}\n"
                            if version.get("endDate"):
                                text += f"  End Date: {version.get('endDate')}\n"

                            if (
                                "_embedded" in version
                                and "definingProject" in version["_embedded"]
                            ):
                                text += f"  Project: {version['_embedded']['definingProject'].get('name', 'Unknown')}\n"

                            if version.get("description", {}).get("raw"):
                                text += (
                                    f"  Description: {version['description']['raw']}\n"
                                )
                            text += "\n"

                    return [TextContent(type="text", text=text)]

                elif name == "create_version":
                    project_id = arguments["project_id"]

                    data = {"name": arguments["name"]}

                    # Add optional fields
                    for field in ["description", "start_date", "end_date", "status"]:
                        if field in arguments:
                            data[field] = arguments[field]

                    result = await self.client.create_version(project_id, data)

                    text = f"✅ Version created successfully:\n\n"
                    text += f"- **Name**: {result.get('name', 'N/A')}\n"
                    text += f"- **ID**: {result.get('id', 'N/A')}\n"
                    text += f"- **Status**: {result.get('status', 'Unknown')}\n"

                    if result.get("startDate"):
                        text += f"- **Start Date**: {result.get('startDate')}\n"
                    if result.get("endDate"):
                        text += f"- **End Date**: {result.get('endDate')}\n"

                    if (
                        "_embedded" in result
                        and "definingProject" in result["_embedded"]
                    ):
                        text += f"- **Project**: {result['_embedded']['definingProject'].get('name', 'Unknown')}\n"

                    return [TextContent(type="text", text=text)]

                elif name == "check_permissions":
                    user_info = await self.client.check_permissions()

                    if not user_info:
                        text = "❌ Unable to retrieve user permissions."
                    else:
                        text = f"**Current User Permissions:**\n\n"
                        text += f"- **Name**: {user_info.get('name', 'Unknown')}\n"
                        text += f"- **ID**: {user_info.get('id', 'N/A')}\n"
                        text += f"- **Email**: {user_info.get('email', 'N/A')}\n"
                        text += f"- **Status**: {user_info.get('status', 'Unknown')}\n"
                        text += f"- **Administrator**: {'Yes' if user_info.get('admin') else 'No'}\n"
                        text += f"- **Language**: {user_info.get('language', 'N/A')}\n"
                        text += f"- **Created**: {user_info.get('createdAt', 'N/A')}\n"

                        # Check for specific permission-related links
                        if "_links" in user_info:
                            links = user_info["_links"]
                            text += f"\n**Available Actions:**\n"
                            for link_name, link_info in links.items():
                                if link_name not in ["self", "showUser"]:
                                    text += f"- {link_name}: Available\n"

                        text += f"\n**Tip**: Use this information to understand why certain operations may fail due to insufficient permissions."

                    return [TextContent(type="text", text=text)]

                elif name == "create_project":
                    data = {
                        "name": arguments["name"],
                        "identifier": arguments["identifier"],
                    }

                    # Add optional fields
                    for field in ["description", "public", "status", "parent_id"]:
                        if field in arguments:
                            data[field] = arguments[field]

                    result = await self.client.create_project(data)

                    text = f"✅ Project created successfully:\n\n"
                    text += f"- **Name**: {result.get('name', 'N/A')}\n"
                    text += f"- **ID**: #{result.get('id', 'N/A')}\n"
                    text += f"- **Identifier**: {result.get('identifier', 'N/A')}\n"
                    text += f"- **Public**: {'Yes' if result.get('public') else 'No'}\n"
                    text += f"- **Status**: {result.get('status', 'N/A')}\n"

                    return [TextContent(type="text", text=text)]

                elif name == "update_project":
                    project_id = arguments["project_id"]

                    # Prepare update data
                    update_data = {}
                    for field in [
                        "name",
                        "identifier",
                        "description",
                        "public",
                        "status",
                        "parent_id",
                    ]:
                        if field in arguments:
                            update_data[field] = arguments[field]

                    if not update_data:
                        return [
                            TextContent(
                                type="text", text="❌ No fields provided to update."
                            )
                        ]

                    result = await self.client.update_project(project_id, update_data)

                    text = f"✅ Project #{project_id} updated successfully:\n\n"
                    text += f"- **Name**: {result.get('name', 'N/A')}\n"
                    text += f"- **Identifier**: {result.get('identifier', 'N/A')}\n"
                    text += f"- **Public**: {'Yes' if result.get('public') else 'No'}\n"
                    text += f"- **Status**: {result.get('status', 'N/A')}\n"

                    return [TextContent(type="text", text=text)]

                elif name == "delete_project":
                    project_id = arguments["project_id"]

                    success = await self.client.delete_project(project_id)

                    if success:
                        text = f"✅ Project #{project_id} deleted successfully."
                    else:
                        text = f"❌ Failed to delete project #{project_id}."

                    return [TextContent(type="text", text=text)]

                elif name == "get_project":
                    project_id = arguments["project_id"]
                    result = await self.client.get_project(project_id)

                    text = f"**Project Details:**\n\n"
                    text += f"- **Name**: {result.get('name', 'N/A')}\n"
                    text += f"- **ID**: #{result.get('id', 'N/A')}\n"
                    text += f"- **Identifier**: {result.get('identifier', 'N/A')}\n"
                    text += f"- **Description**: {result.get('description', {}).get('raw', 'No description') if result.get('description') else 'No description'}\n"
                    text += f"- **Public**: {'Yes' if result.get('public') else 'No'}\n"
                    text += f"- **Status**: {result.get('status', 'N/A')}\n"
                    text += f"- **Created**: {result.get('createdAt', 'N/A')}\n"
                    text += f"- **Updated**: {result.get('updatedAt', 'N/A')}\n"

                    return [TextContent(type="text", text=text)]

                elif name == "create_membership":
                    data = {"project_id": arguments["project_id"]}

                    # Add user or group
                    if "user_id" in arguments:
                        data["user_id"] = arguments["user_id"]
                    elif "group_id" in arguments:
                        data["group_id"] = arguments["group_id"]
                    else:
                        return [
                            TextContent(
                                type="text",
                                text="❌ Either user_id or group_id is required.",
                            )
                        ]

                    # Add roles
                    if "role_ids" in arguments:
                        data["role_ids"] = arguments["role_ids"]
                    elif "role_id" in arguments:
                        data["role_id"] = arguments["role_id"]
                    else:
                        return [
                            TextContent(
                                type="text",
                                text="❌ Either role_ids or role_id is required.",
                            )
                        ]

                    # Add optional fields
                    if "notification_message" in arguments:
                        data["notification_message"] = arguments["notification_message"]

                    result = await self.client.create_membership(data)

                    text = f"✅ Membership created successfully:\n\n"
                    text += f"- **ID**: #{result.get('id', 'N/A')}\n"

                    if "_embedded" in result:
                        embedded = result["_embedded"]
                        if "project" in embedded:
                            text += f"- **Project**: {embedded['project'].get('name', 'Unknown')}\n"
                        if "principal" in embedded:
                            text += f"- **User/Group**: {embedded['principal'].get('name', 'Unknown')}\n"
                        if "roles" in embedded:
                            roles = [
                                role.get("name", "Unknown")
                                for role in embedded["roles"]
                            ]
                            text += f"- **Roles**: {', '.join(roles)}\n"

                    return [TextContent(type="text", text=text)]

                elif name == "update_membership":
                    membership_id = arguments["membership_id"]

                    # Prepare update data
                    update_data = {}
                    if "role_ids" in arguments:
                        update_data["role_ids"] = arguments["role_ids"]
                    elif "role_id" in arguments:
                        update_data["role_id"] = arguments["role_id"]

                    if "notification_message" in arguments:
                        update_data["notification_message"] = arguments[
                            "notification_message"
                        ]

                    if not update_data:
                        return [
                            TextContent(
                                type="text", text="❌ No fields provided to update."
                            )
                        ]

                    result = await self.client.update_membership(
                        membership_id, update_data
                    )

                    text = f"✅ Membership #{membership_id} updated successfully:\n\n"

                    if "_embedded" in result:
                        embedded = result["_embedded"]
                        if "roles" in embedded:
                            roles = [
                                role.get("name", "Unknown")
                                for role in embedded["roles"]
                            ]
                            text += f"- **Roles**: {', '.join(roles)}\n"

                    return [TextContent(type="text", text=text)]

                elif name == "delete_membership":
                    membership_id = arguments["membership_id"]

                    success = await self.client.delete_membership(membership_id)

                    if success:
                        text = f"✅ Membership #{membership_id} deleted successfully."
                    else:
                        text = f"❌ Failed to delete membership #{membership_id}."

                    return [TextContent(type="text", text=text)]

                elif name == "get_membership":
                    membership_id = arguments["membership_id"]
                    result = await self.client.get_membership(membership_id)

                    text = f"**Membership Details:**\n\n"
                    text += f"- **ID**: #{result.get('id', 'N/A')}\n"

                    if "_embedded" in result:
                        embedded = result["_embedded"]
                        if "project" in embedded:
                            text += f"- **Project**: {embedded['project'].get('name', 'Unknown')}\n"
                        if "principal" in embedded:
                            text += f"- **User/Group**: {embedded['principal'].get('name', 'Unknown')}\n"
                        if "roles" in embedded:
                            roles = [
                                role.get("name", "Unknown")
                                for role in embedded["roles"]
                            ]
                            text += f"- **Roles**: {', '.join(roles)}\n"

                    return [TextContent(type="text", text=text)]

                elif name == "list_project_members":
                    project_id = arguments["project_id"]

                    # Filter memberships by project
                    filters = json.dumps(
                        [{"project": {"operator": "=", "values": [str(project_id)]}}]
                    )
                    result = await self.client.get_memberships(project_id=project_id)
                    memberships = result.get("_embedded", {}).get("elements", [])

                    if not memberships:
                        text = f"No members found for project #{project_id}."
                    else:
                        text = f"**Project #{project_id} Members ({len(memberships)}):**\n\n"
                        for membership in memberships:
                            if "_embedded" in membership:
                                embedded = membership["_embedded"]
                                user_name = "Unknown"
                                roles = []

                                if "principal" in embedded:
                                    user_name = embedded["principal"].get(
                                        "name", "Unknown"
                                    )
                                if "roles" in embedded:
                                    roles = [
                                        role.get("name", "Unknown")
                                        for role in embedded["roles"]
                                    ]

                                text += f"- **{user_name}**: {', '.join(roles)}\n"

                    return [TextContent(type="text", text=text)]

                elif name == "list_user_projects":
                    user_id = arguments["user_id"]

                    # Filter memberships by user
                    result = await self.client.get_memberships(user_id=user_id)
                    memberships = result.get("_embedded", {}).get("elements", [])

                    if not memberships:
                        text = f"No projects found for user #{user_id}."
                    else:
                        text = f"**User #{user_id} Projects ({len(memberships)}):**\n\n"
                        for membership in memberships:
                            if "_embedded" in membership:
                                embedded = membership["_embedded"]
                                project_name = "Unknown"
                                roles = []

                                if "project" in embedded:
                                    project_name = embedded["project"].get(
                                        "name", "Unknown"
                                    )
                                if "roles" in embedded:
                                    roles = [
                                        role.get("name", "Unknown")
                                        for role in embedded["roles"]
                                    ]

                                text += f"- **{project_name}**: {', '.join(roles)}\n"

                    return [TextContent(type="text", text=text)]

                elif name == "list_roles":
                    result = await self.client.get_roles()
                    roles = result.get("_embedded", {}).get("elements", [])

                    if not roles:
                        text = "No roles found."
                    else:
                        text = f"Available roles ({len(roles)}):\n\n"
                        for role in roles:
                            text += f"- **{role.get('name', 'Unnamed')}** (ID: {role.get('id', 'N/A')})\n"

                    return [TextContent(type="text", text=text)]

                elif name == "get_role":
                    role_id = arguments["role_id"]
                    result = await self.client.get_role(role_id)

                    text = f"**Role Details:**\n\n"
                    text += f"- **Name**: {result.get('name', 'N/A')}\n"
                    text += f"- **ID**: #{result.get('id', 'N/A')}\n"

                    # Add any additional role information if available
                    if "permissions" in result:
                        permissions = result["permissions"]
                        if permissions:
                            text += f"- **Permissions**: {len(permissions)} permissions assigned\n"

                    return [TextContent(type="text", text=text)]

                elif name == "set_work_package_parent":
                    work_package_id = arguments["work_package_id"]
                    parent_id = arguments["parent_id"]

                    result = await self.client.set_work_package_parent(
                        work_package_id, parent_id
                    )

                    text = f"✅ Parent relationship created successfully:\n\n"
                    text += f"- **Child Work Package**: #{work_package_id}\n"
                    text += f"- **Parent Work Package**: #{parent_id}\n"
                    text += f"- **Subject**: {result.get('subject', 'N/A')}\n"

                    if "_links" in result and "parent" in result["_links"]:
                        parent_href = result["_links"]["parent"].get("href", "")
                        text += f"- **Parent Link**: {parent_href}\n"

                    return [TextContent(type="text", text=text)]

                elif name == "remove_work_package_parent":
                    work_package_id = arguments["work_package_id"]

                    result = await self.client.remove_work_package_parent(
                        work_package_id
                    )

                    text = f"✅ Parent relationship removed successfully:\n\n"
                    text += f"- **Work Package**: #{work_package_id} is now top-level\n"
                    text += f"- **Subject**: {result.get('subject', 'N/A')}\n"

                    return [TextContent(type="text", text=text)]

                elif name == "list_work_package_children":
                    parent_id = arguments["parent_id"]
                    include_descendants = arguments.get("include_descendants", False)

                    result = await self.client.list_work_package_children(
                        parent_id, include_descendants
                    )
                    children = result.get("_embedded", {}).get("elements", [])

                    if not children:
                        text = f"No {'descendants' if include_descendants else 'children'} found for work package #{parent_id}."
                    else:
                        text = f"**{'Descendants' if include_descendants else 'Children'} of Work Package #{parent_id} ({len(children)}):**\n\n"
                        for child in children:
                            text += f"- **#{child.get('id', 'N/A')}**: {child.get('subject', 'No subject')}\n"

                            # Show type and status if available
                            if "_embedded" in child:
                                embedded = child["_embedded"]
                                if "type" in embedded:
                                    text += f"  Type: {embedded['type'].get('name', 'Unknown')}\n"
                                if "status" in embedded:
                                    text += f"  Status: {embedded['status'].get('name', 'Unknown')}\n"
                            text += "\n"

                    return [TextContent(type="text", text=text)]

                elif name == "create_work_package_relation":
                    data = {
                        "from_id": arguments["from_id"],
                        "to_id": arguments["to_id"],
                        "relation_type": arguments["relation_type"],
                    }

                    # Add optional fields
                    for field in ["lag", "description"]:
                        if field in arguments:
                            data[field] = arguments[field]

                    result = await self.client.create_work_package_relation(data)

                    text = f"✅ Work package relation created successfully:\n\n"
                    text += f"- **Relation ID**: #{result.get('id', 'N/A')}\n"
                    text += f"- **Type**: {result.get('type', 'N/A')}\n"
                    text += f"- **From**: Work Package #{arguments['from_id']}\n"
                    text += f"- **To**: Work Package #{arguments['to_id']}\n"

                    if "lag" in result:
                        text += f"- **Lag**: {result.get('lag', 0)} working days\n"
                    if "description" in result:
                        text += (
                            f"- **Description**: {result.get('description', 'N/A')}\n"
                        )

                    return [TextContent(type="text", text=text)]

                elif name == "list_work_package_relations":
                    filters = None
                    filter_conditions = []

                    if "work_package_id" in arguments:
                        wp_id = arguments["work_package_id"]
                        filter_conditions.append(
                            {"involved": {"operator": "=", "values": [str(wp_id)]}}
                        )

                    if "relation_type" in arguments:
                        rel_type = arguments["relation_type"]
                        filter_conditions.append(
                            {"type": {"operator": "=", "values": [rel_type]}}
                        )

                    if filter_conditions:
                        filters = json.dumps(filter_conditions)

                    result = await self.client.list_work_package_relations(filters)
                    relations = result.get("_embedded", {}).get("elements", [])

                    if not relations:
                        text = "No work package relations found."
                    else:
                        text = f"**Work Package Relations ({len(relations)}):**\n\n"
                        for relation in relations:
                            text += f"- **#{relation.get('id', 'N/A')}**: {relation.get('type', 'Unknown')} relation\n"

                            if "_embedded" in relation:
                                embedded = relation["_embedded"]
                                if "from" in embedded and "to" in embedded:
                                    from_wp = embedded["from"]
                                    to_wp = embedded["to"]
                                    text += f"  From: #{from_wp.get('id', 'N/A')} - {from_wp.get('subject', 'No subject')}\n"
                                    text += f"  To: #{to_wp.get('id', 'N/A')} - {to_wp.get('subject', 'No subject')}\n"

                            if "lag" in relation:
                                text += (
                                    f"  Lag: {relation.get('lag', 0)} working days\n"
                                )
                            if "description" in relation:
                                text += f"  Description: {relation.get('description', 'N/A')}\n"
                            text += "\n"

                    return [TextContent(type="text", text=text)]

                elif name == "update_work_package_relation":
                    relation_id = arguments["relation_id"]

                    # Prepare update data
                    update_data = {}
                    for field in ["relation_type", "lag", "description"]:
                        if field in arguments:
                            update_data[field] = arguments[field]

                    if not update_data:
                        return [
                            TextContent(
                                type="text", text="❌ No fields provided to update."
                            )
                        ]

                    result = await self.client.update_work_package_relation(
                        relation_id, update_data
                    )

                    text = f"✅ Work package relation #{relation_id} updated successfully:\n\n"
                    text += f"- **Type**: {result.get('type', 'N/A')}\n"

                    if "lag" in result:
                        text += f"- **Lag**: {result.get('lag', 0)} working days\n"
                    if "description" in result:
                        text += (
                            f"- **Description**: {result.get('description', 'N/A')}\n"
                        )

                    return [TextContent(type="text", text=text)]

                elif name == "delete_work_package_relation":
                    relation_id = arguments["relation_id"]

                    success = await self.client.delete_work_package_relation(
                        relation_id
                    )

                    if success:
                        text = f"✅ Work package relation #{relation_id} deleted successfully."
                    else:
                        text = (
                            f"❌ Failed to delete work package relation #{relation_id}."
                        )

                    return [TextContent(type="text", text=text)]

                elif name == "get_work_package_relation":
                    relation_id = arguments["relation_id"]
                    result = await self.client.get_work_package_relation(relation_id)

                    text = f"**Work Package Relation Details:**\n\n"
                    text += f"- **ID**: #{result.get('id', 'N/A')}\n"
                    text += f"- **Type**: {result.get('type', 'N/A')}\n"
                    text += f"- **Reverse Type**: {result.get('reverseType', 'N/A')}\n"

                    if "_embedded" in result:
                        embedded = result["_embedded"]
                        if "from" in embedded and "to" in embedded:
                            from_wp = embedded["from"]
                            to_wp = embedded["to"]
                            text += f"- **From**: #{from_wp.get('id', 'N/A')} - {from_wp.get('subject', 'No subject')}\n"
                            text += f"- **To**: #{to_wp.get('id', 'N/A')} - {to_wp.get('subject', 'No subject')}\n"

                    if "lag" in result:
                        text += f"- **Lag**: {result.get('lag', 0)} working days\n"
                    if "description" in result:
                        text += (
                            f"- **Description**: {result.get('description', 'N/A')}\n"
                        )

                    return [TextContent(type="text", text=text)]

                else:
                    return [TextContent(type="text", text=f"Unknown tool: {name}")]

            except Exception as e:
                logger.error(f"Error executing tool {name}: {e}", exc_info=True)

                error_text = f"❌ Error executing tool '{name}':\n\n{str(e)}"

                return [TextContent(type="text", text=error_text)]

    async def run(self):
        """Start the MCP server"""
        # Initialize OpenProject client from environment variables
        base_url = os.getenv("OPENPROJECT_URL")
        api_key = os.getenv("OPENPROJECT_API_KEY")
        proxy = os.getenv("OPENPROJECT_PROXY")  # Optional proxy

        if not base_url or not api_key:
            logger.error("OPENPROJECT_URL or OPENPROJECT_API_KEY not set!")
            logger.info("Please set the required environment variables in .env file")
        else:
            self.client = OpenProjectClient(base_url, api_key, proxy)
            logger.info(f"✅ OpenProject Client initialized for {base_url}")

            # Optional: Test connection on startup
            if os.getenv("TEST_CONNECTION_ON_STARTUP", "false").lower() == "true":
                try:
                    await self.client.test_connection()
                    logger.info("✅ API connection test successful!")
                except Exception as e:
                    logger.error(f"❌ API connection test failed: {e}")

        # Start the server
        from mcp.server.stdio import stdio_server

        async with stdio_server() as (read_stream, write_stream):
            await self.server.run(
                read_stream, write_stream, self.server.create_initialization_options()
            )


async def main():
    """Main entry point"""
    logger.info(f"Starting OpenProject MCP Server v{__version__}")

    server = OpenProjectMCPServer()
    await server.run()


if __name__ == "__main__":
    asyncio.run(main())
